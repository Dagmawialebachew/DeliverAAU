# database/db.py (Postgres/asyncpg migration)
from collections import Counter
from decimal import Decimal
import json
import os
import random
import asyncpg
from config import settings
from typing import Optional, Dict, Any, List, Tuple
import datetime
from asyncpg.connection import Connection
from asyncpg.pool import Pool
    
from datetime import datetime, timedelta, date
from typing import Dict, Any, List, Tuple, Optional
from decimal import Decimal
import math

# reportlab import for PDF export (install with pip install reportlab)
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from datetime import datetime
# --- 1. UNIFIED SCHEMA SQL (Postgres Dialect) ---
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS users (
    id SERIAL PRIMARY KEY,
    telegram_id BIGINT UNIQUE,
role TEXT,
    first_name TEXT,
    phone TEXT,
    campus TEXT,
    coins INTEGER DEFAULT 0,
    xp INTEGER DEFAULT 0,
    level INTEGER DEFAULT 1,
    status TEXT DEFAULT 'active'
);


CREATE TABLE IF NOT EXISTS tickets (
    ticket_id TEXT PRIMARY KEY,
    user_id BIGINT NOT NULL,
    text TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'open',
    original_msg_id BIGINT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    closed_at TIMESTAMP
);


CREATE TABLE IF NOT EXISTS delivery_guys (
    id SERIAL PRIMARY KEY,
    user_id BIGINT UNIQUE, -- Foreign key to users.id
    telegram_id BIGINT UNIQUE, -- Telegram chat_id for messaging
    name TEXT,
    campus TEXT,
    phone TEXT, 
    active BOOLEAN DEFAULT FALSE, -- Changed INTEGER DEFAULT 0 to BOOLEAN DEFAULT FALSE
    blocked BOOLEAN DEFAULT FALSE, -- Changed INTEGER DEFAULT 0 to BOOLEAN DEFAULT FALSE
    total_deliveries INTEGER DEFAULT 0,
    accepted_requests INTEGER DEFAULT 0,
    total_requests INTEGER DEFAULT 0,
    last_lat DOUBLE PRECISION NULL, -- Changed REAL to DOUBLE PRECISION
    last_lon DOUBLE PRECISION NULL, -- Changed REAL to DOUBLE PRECISION
    skipped_requests INTEGER DEFAULT 0,
    last_skip_at TIMESTAMP NULL,
    last_online_at TIMESTAMP NULL,
    last_offline_at TIMESTAMP NULL,
    coins INTEGER DEFAULT 0,
    xp INTEGER DEFAULT 0,
    level INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS vendors (
    id SERIAL PRIMARY KEY,
    telegram_id BIGINT UNIQUE, 
    name TEXT NOT NULL, 
    menu_json TEXT, 
    status TEXT DEFAULT 'active',
    rating_avg DOUBLE PRECISION DEFAULT 0.0, -- Changed REAL to DOUBLE PRECISION
    rating_count INTEGER DEFAULT 0,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);



CREATE TABLE IF NOT EXISTS orders (
    id SERIAL PRIMARY KEY,
    user_id INTEGER,
    delivery_guy_id INTEGER NULL, 
    vendor_id INTEGER,
    pickup TEXT,
    dropoff TEXT,
    items_json TEXT,
    food_subtotal DOUBLE PRECISION DEFAULT 0.0, -- Changed REAL to DOUBLE PRECISION
    delivery_fee DOUBLE PRECISION DEFAULT 0.0, -- Changed REAL to DOUBLE PRECISION
    status TEXT, -- pending / assigned / preparing / ready / in_progress / delivered / cancelled
    payment_method TEXT,
    payment_status TEXT,
    receipt_id INTEGER,
    breakdown_json TEXT,
    live_shared BOOLEAN DEFAULT FALSE, -- Changed INTEGER DEFAULT 0 to BOOLEAN DEFAULT FALSE
    live_expires TIMESTAMP NULL, 
    last_lat DOUBLE PRECISION NULL, -- Changed REAL to DOUBLE PRECISION
    last_lon DOUBLE PRECISION NULL, -- Changed REAL to DOUBLE PRECISION
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    accepted_at TIMESTAMP NULL,
    delivered_at TIMESTAMP NULL,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


CREATE TABLE IF NOT EXISTS daily_stats (
    id SERIAL PRIMARY KEY,
    dg_id INTEGER,
    date TEXT, -- YYYY-MM-DD
    deliveries INTEGER DEFAULT 0,
    earnings DOUBLE PRECISION DEFAULT 0.0, -- Changed REAL to DOUBLE PRECISION
    skipped INTEGER DEFAULT 0,
    assigned INTEGER DEFAULT 0, 
    acceptance_rate DOUBLE PRECISION DEFAULT 0.0, -- Changed REAL to DOUBLE PRECISION
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(dg_id, date)
);


CREATE TABLE IF NOT EXISTS ratings (
    id SERIAL PRIMARY KEY,
    order_id INTEGER NOT NULL,
    delivery_guy_id INTEGER,
    vendor_id INTEGER,
    stars INTEGER NOT NULL,
    comment TEXT,
    type TEXT NOT NULL CHECK (type IN ('delivery','vendor')),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    CONSTRAINT unique_order_type UNIQUE (order_id, type)
);



CREATE TABLE IF NOT EXISTS media (
    id SERIAL PRIMARY KEY,
    file_path TEXT,
    uploader_user_id INTEGER,
    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    verified_by INTEGER,
    verify_status TEXT,
    notes TEXT
);


CREATE TABLE IF NOT EXISTS daily_stats_archive (
    id SERIAL PRIMARY KEY,
    dg_id INTEGER NOT NULL,
    date DATE NOT NULL,
    deliveries INTEGER DEFAULT 0,
    earnings DOUBLE PRECISION DEFAULT 0.0,
    skipped INTEGER DEFAULT 0,
    assigned INTEGER DEFAULT 0,
    acceptance_rate DOUBLE PRECISION DEFAULT 0.0,
    archived_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(dg_id, date)
);


CREATE TABLE IF NOT EXISTS subscriptions (
    id SERIAL PRIMARY KEY,
    user_id INTEGER,
    type TEXT,
    start_ts TIMESTAMP,
    end_ts TIMESTAMP,
    status TEXT
);

CREATE TABLE IF NOT EXISTS location_logs (
    id SERIAL PRIMARY KEY,
    order_id INTEGER NULL, 
    delivery_guy_id INTEGER NULL, 
    lat DOUBLE PRECISION, -- Changed REAL to DOUBLE PRECISION
    lon DOUBLE PRECISION, -- Changed REAL to DOUBLE PRECISION
    ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


CREATE TABLE IF NOT EXISTS admin_settings (
    key TEXT PRIMARY KEY,
    value_json TEXT
);

CREATE TABLE IF NOT EXISTS jobs_log (
    id SERIAL PRIMARY KEY,
    job_name TEXT,
    key TEXT,
    status TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


CREATE TABLE IF NOT EXISTS leaderboards (
    user_id BIGINT PRIMARY KEY,           -- same as orders.user_id (telegram_id)
    display_name TEXT NOT NULL,           -- current Telegram name snapshot
    bites INTEGER DEFAULT 0,              -- points for leaderboard
    rank INTEGER DEFAULT NULL,            -- cached computed rank (optional)
    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS asbeza_items (
    id SERIAL PRIMARY KEY,
    name TEXT NOT NULL,
    description TEXT,
    base_price DOUBLE PRECISION NOT NULL,
    image_url TEXT,
    active BOOLEAN DEFAULT TRUE,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS asbeza_variants (
    id SERIAL PRIMARY KEY,
    item_id INTEGER REFERENCES asbeza_items(id) ON DELETE CASCADE,
    name TEXT NOT NULL,              -- e.g. "Mint", "Strawberry", "Blue"
    price DOUBLE PRECISION NOT NULL, -- can override base price
    stock INTEGER DEFAULT 0
);


CREATE TABLE IF NOT EXISTS asbeza_orders (
    id SERIAL PRIMARY KEY,
    user_id BIGINT,
    total_price DOUBLE PRECISION,
    upfront_paid DOUBLE PRECISION,
    status TEXT, -- pending / confirmed / delivered / cancelled
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS asbeza_order_items (
    id SERIAL PRIMARY KEY,
    order_id INTEGER REFERENCES asbeza_orders(id) ON DELETE CASCADE,
    variant_id INTEGER REFERENCES asbeza_variants(id),
    quantity INTEGER,
    price DOUBLE PRECISION
);

-- optional: store payment attempts / history separately (recommended)
CREATE TABLE IF NOT EXISTS asbeza_order_payments (
  id SERIAL PRIMARY KEY,
  order_id INTEGER REFERENCES asbeza_orders(id) ON DELETE CASCADE,
  user_id BIGINT,
  amount DOUBLE PRECISION,
  payment_proof_url TEXT,
  method TEXT,
  status TEXT DEFAULT 'pending', -- pending / verified / rejected
  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


CREATE TABLE IF NOT EXISTS admins (
    id SERIAL PRIMARY KEY,
    username VARCHAR(50) UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role VARCHAR(20) DEFAULT 'admin',
    created_at TIMESTAMP DEFAULT NOW()
);

ALTER TABLE asbeza_variants
ADD COLUMN IF NOT EXISTS image_url TEXT,
ADD COLUMN IF NOT EXISTS cost_price DOUBLE PRECISION DEFAULT 0;

ALTER TABLE asbeza_items 
ADD COLUMN IF NOT EXISTS category TEXT;





-- Helpful indexes
CREATE INDEX IF NOT EXISTS idx_leaderboards_bites ON leaderboards(bites DESC);
CREATE INDEX IF NOT EXISTS idx_leaderboards_updated ON leaderboards(last_updated);

-- This helps the DB find a user's orders instantly
CREATE INDEX IF NOT EXISTS idx_asbeza_orders_user_id ON asbeza_orders(user_id);

-- This helps the DB count items instantly
CREATE INDEX IF NOT EXISTS idx_asbeza_order_items_order_id ON asbeza_order_items(order_id);

-- This helps the DB find payments instantly
CREATE INDEX IF NOT EXISTS idx_asbeza_order_payments_order_id ON asbeza_order_payments(order_id);


-- Indexes (Postgres doesn't need IF NOT EXISTS for CREATE INDEX in a transaction block
-- but this is fine for initial setup)
CREATE INDEX IF NOT EXISTS idx_users_telegram_id ON users(telegram_id);
CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
CREATE INDEX IF NOT EXISTS idx_orders_status_updated ON orders(status, updated_at);
CREATE INDEX IF NOT EXISTS idx_orders_created ON orders(created_at);
CREATE INDEX IF NOT EXISTS idx_subscriptions_user_end ON subscriptions(user_id, end_ts);
CREATE INDEX IF NOT EXISTS idx_location_order_ts ON location_logs(order_id, ts);
CREATE INDEX IF NOT EXISTS idx_daily_stats_dg_date ON daily_stats(dg_id, date);
CREATE INDEX IF NOT EXISTS idx_vendors_telegram_id ON vendors(telegram_id);
CREATE INDEX IF NOT EXISTS idx_vendors_name ON vendors(name);
ALTER TABLE orders ALTER COLUMN user_id TYPE BIGINT;
ALTER TABLE orders ALTER COLUMN delivery_guy_id TYPE BIGINT;
ALTER TABLE orders ADD COLUMN IF NOT EXISTS notes TEXT;
ALTER TABLE orders ALTER COLUMN vendor_id TYPE BIGINT;
ALTER TABLE users ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
ALTER TABLE users ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
ALTER TABLE orders ADD COLUMN IF NOT EXISTS ready_at TIMESTAMP NULL;


ALTER TABLE users
ADD COLUMN IF NOT EXISTS gender TEXT CHECK (gender IN ('male','female'));


ALTER TABLE asbeza_orders 
ADD COLUMN IF NOT EXISTS delivery_fee DOUBLE PRECISION DEFAULT 0,
ADD COLUMN IF NOT EXISTS payment_proof_url TEXT, 
ADD COLUMN IF NOT EXISTS created_by_ip TEXT,
ADD COLUMN IF NOT EXISTS delivery_guy_id INTEGER REFERENCES delivery_guys(id),
ADD COLUMN IF NOT EXISTS delivered_at TIMESTAMP NULL;
-- Add delivered_at to measure fulfillment speed
ALTER TABLE asbeza_orders
ADD COLUMN IF NOT EXISTS delivered_at TIMESTAMP;

-- Add indexes for analytics queries
CREATE INDEX IF NOT EXISTS idx_orders_created_at ON asbeza_orders (created_at);
CREATE INDEX IF NOT EXISTS idx_orders_status ON asbeza_orders (status);
CREATE INDEX IF NOT EXISTS idx_orders_user_id ON asbeza_orders (user_id);
CREATE INDEX IF NOT EXISTS idx_order_items_variant_id ON asbeza_order_items (variant_id);
CREATE INDEX IF NOT EXISTS idx_variants_stock ON asbeza_variants (stock);
CREATE INDEX IF NOT EXISTS idx_payments_order_id ON asbeza_order_payments (order_id);
CREATE INDEX IF NOT EXISTS idx_orders_delivery_guy ON asbeza_orders (delivery_guy_id);


ALTER TABLE delivery_guys
ADD COLUMN IF NOT EXISTS gender TEXT CHECK (gender IN ('male','female'));



-- Make sure dg_id is BIGINT
ALTER TABLE daily_stats ALTER COLUMN dg_id TYPE BIGINT;
ALTER TABLE daily_stats ALTER COLUMN date TYPE DATE USING date::date;



ALTER TABLE orders
ADD COLUMN IF NOT EXISTS expires_at TIMESTAMP NULL,
ADD COLUMN IF NOT EXISTS vendor_confirmed_at TIMESTAMP NULL,
ADD COLUMN IF NOT EXISTS cancel_reason TEXT NULL;


-- Optional: index to speed up expiry checks
CREATE INDEX IF NOT EXISTS idx_orders_expires ON orders(expires_at);
CREATE INDEX IF NOT EXISTS idx_orders_vendor_created ON orders(vendor_id, created_at);
CREATE INDEX IF NOT EXISTS idx_orders_pending_created
ON orders(created_at)
WHERE status = 'pending';


CREATE INDEX IF NOT EXISTS idx_orders_user_id_created ON orders(user_id, created_at);
CREATE INDEX IF NOT EXISTS idx_orders_dg_id_created ON orders(delivery_guy_id, created_at);
CREATE INDEX IF NOT EXISTS idx_ratings_vendor_id ON ratings(vendor_id);
CREATE INDEX IF NOT EXISTS idx_ratings_dg_id ON ratings(delivery_guy_id);

-- 1) Users: referral + combo flag
ALTER TABLE users
ADD COLUMN IF NOT EXISTS referral_code TEXT UNIQUE,
ADD COLUMN IF NOT EXISTS referred_by INTEGER NULL,
ADD COLUMN IF NOT EXISTS genna_combo_unlocked BOOLEAN DEFAULT FALSE;

-- 2) Spin entries: isolated store keyed by BIGINT user_id (aligns with orders.user_id)
CREATE TABLE IF NOT EXISTS spin_entries (
    user_id BIGINT PRIMARY KEY,
    total_entries INTEGER DEFAULT 0,
    available_spins INTEGER DEFAULT 0,
    last_spin_date TIMESTAMP NULL
);

-- 3) Helpful indexes
CREATE INDEX IF NOT EXISTS idx_orders_user_status ON orders(user_id, status);

-- Recompute spin_entries.available_spins for all users (run nightly)
UPDATE spin_entries s
SET available_spins = GREATEST((l.bites / 25)::int - s.total_entries, 0)
FROM leaderboards l
WHERE s.user_id = l.user_id;


CREATE TABLE IF NOT EXISTS spin_rewards (
    id SERIAL PRIMARY KEY,
    user_id BIGINT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
    reward JSONB NOT NULL,                -- stores prize info as JSON
    claimed BOOLEAN DEFAULT FALSE,        -- whether reward has been redeemed
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


-- Snapshot weekly leaderboard (optional)
CREATE TABLE IF NOT EXISTS weekly_snapshots (
  week_start DATE PRIMARY KEY,
  snapshot_json JSONB,
  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);


"""

class Database:
    def __init__(self):
        self.database_url = os.environ.get("DB_PATH")  # use DATABASE_URL not DB_PATH
        if not self.database_url:
            raise ValueError("DATABASE_URL environment variable is not set.")
        self._pool: Optional[Pool] = None

    async def init_pool(self):
        if self._pool:
            await self._pool.close()
        self._pool = await asyncpg.create_pool(
            self.database_url,
            min_size=1,
            max_size=20,
            statement_cache_size=0  # 🔥 THIS LINE FIXES IT
        )
    async def reset_schema(self):
        """Drop all data and recreate schema fresh."""
        async with self._open_connection() as conn:
            # Drop tables if they exist
            await conn.execute("""
                DROP TABLE IF EXISTS 
                    users, tickets, delivery_guys, vendors, orders,
                    daily_stats, ratings, media, daily_stats_archive,
                    subscriptions, location_logs, admin_settings, jobs_log
                CASCADE;
            """)
            # Recreate schema
            await conn.execute(SCHEMA_SQL)

    def _get_pool(self) -> Pool:
        """Return the pool synchronously (must be initialized first)."""
        if not self._pool:
            raise RuntimeError("Pool not initialized. Call init_pool() first.")
        return self._pool
    
    async def close_pool(self):
        if self._pool:
            await self._pool.close()
            self._pool = None
    
    async def init_schema(self):
        """Run SCHEMA_SQL to create tables if they don't exist."""
        async with self._open_connection() as conn:
            await conn.execute(SCHEMA_SQL)

    def _open_connection(self):
        """Return an async context manager for acquiring a connection."""
        return self._get_pool().acquire()

    @staticmethod
    def _row_to_dict(record: asyncpg.Record) -> Dict[str, Any]:
        return dict(record)
    
    async def execute(self, query: str, *args):
        """Generic execute for INSERT/UPDATE/DELETE."""
        async with self._open_connection() as conn:
            return await conn.execute(query, *args)

    async def update_user_field(self, telegram_id: int, field: str, value: str):
        """Update a single field in users table."""
        async with self._open_connection() as conn:
            return await conn.execute(
                f"UPDATE users SET {field} = $1 WHERE telegram_id = $2",
                value, telegram_id
            )
  


    # -------------------- Users --------------------
    async def get_user(self, telegram_id: int) -> Optional[Dict[str, Any]]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM users WHERE telegram_id = $1", telegram_id
            )
            return self._row_to_dict(row) if row else None

    async def create_user(
    self,
    telegram_id: int,
    role: str,
    first_name: str,
    phone: str,
    campus: str,
    gender: str = None,
    xp: int = 0,
    referred_by: int | None = None,   # <-- new
) -> int:
        async with self._open_connection() as conn:
            result = await conn.fetchval(
                """
                INSERT INTO users
                (telegram_id, role, first_name, phone, campus, gender, status, xp, referred_by)
                VALUES ($1, $2, $3, $4, $5, $6, 'active', $7, $8)
                ON CONFLICT (telegram_id) DO NOTHING
                RETURNING id
                """,
                telegram_id, role, first_name, phone, campus, gender, xp, referred_by
            )
            if result is None:
                result = await conn.fetchval(
                    "SELECT id FROM users WHERE telegram_id=$1", telegram_id
                )
            return int(result) if result is not None else 0

        
    async def get_user_id_by_telegram(self, telegram_id: int) -> int | None:
        async with self._open_connection() as conn:
            return await conn.fetchval(
                "SELECT id FROM users WHERE telegram_id=$1",
                telegram_id
            )
            
    

    async def get_spin_stats(self, user_id: int) -> dict:
        async with self._open_connection() as conn:
            bites = await conn.fetchval(
                "SELECT bites FROM leaderboards WHERE user_id=$1",
                user_id
            ) or 0

        spins_available = bites // 25
        progress = bites % 25

        return {
            "bites": bites,
            "spins_available": spins_available,
            "progress": progress
        }

    
    async def get_user_bites(self, user_id: int) -> int:
        async with self._open_connection() as conn:
            return await conn.fetchval(
                "SELECT bites FROM leaderboards WHERE user_id=$1",
                user_id
            ) or 0
            

    async def get_user_spins_and_bites(self, user_id: int) -> Tuple[int, int]:
        """
        Returns (available_spins, bites_total).
        available_spins is computed as floor(bites_total/25) - total_entries (spins used).
        """
        async with self._open_connection() as conn:
            bites = await conn.fetchval("SELECT bites FROM leaderboards WHERE user_id=$1", user_id) or 0
            used = await conn.fetchval("SELECT total_entries FROM spin_entries WHERE user_id=$1", user_id) or 0
            total_spins = bites // 25
            available_spins = max(0, total_spins - used)
            progress = bites % 25
            next_threshold = (total_spins + 1) * 25
            remaining = next_threshold - bites
            
            
        return available_spins, bites, progress

    async def get_user_rank_and_progress(self, user_id: int) -> Tuple[Optional[int], float]:
        """
        Returns (rank, progress_pct) where progress_pct is percent to next spin threshold.
        """
        async with self._open_connection() as conn:
            rank = await conn.fetchval(
                """
                SELECT r FROM (
                    SELECT user_id, RANK() OVER (ORDER BY bites DESC) AS r
                    FROM leaderboards
                ) t WHERE user_id=$1
                """,
                user_id
            )
            bites = await conn.fetchval("SELECT bites FROM leaderboards WHERE user_id=$1", user_id) or 0
            next_threshold = ((bites // 25) + 1) * 25
            progress_pct = min(100.0, (bites / next_threshold) * 100) if next_threshold > 0 else 100.0
            return rank, progress_pct

    async def consume_spin(self, user_id: int) -> bool:
        """
        Atomically consume one available spin (increment total_entries, decrement available via logic).
        Returns True if consumed, False if no spins available.
        """
        async with self._open_connection() as conn:
            async with conn.transaction():
                # compute available_spins on DB side to avoid race
                row = await conn.fetchrow(
                    """
                    SELECT
                    (COALESCE(l.bites,0) / 25)::int AS total_spins,
                    COALESCE(s.total_entries,0) AS used_spins
                    FROM (SELECT bites FROM leaderboards WHERE user_id=$1) l
                    LEFT JOIN spin_entries s ON s.user_id=$1
                    """,
                    user_id
                )
                if not row:
                    return False
                total_spins = int(row["total_spins"] or 0)
                used_spins = int(row["used_spins"] or 0)
                available = total_spins - used_spins
                if available <= 0:
                    return False

                # increment used spins and upsert spin_entries
                await conn.execute(
                    """
                    INSERT INTO spin_entries (user_id, total_entries, available_spins, last_spin_date)
                    VALUES ($1, 1, GREATEST($2 - 1, 0), CURRENT_TIMESTAMP)
                    ON CONFLICT (user_id) DO UPDATE
                    SET total_entries = spin_entries.total_entries + 1,
                        available_spins = GREATEST((COALESCE((SELECT bites FROM leaderboards WHERE user_id=$1),0) / 25)::int - (spin_entries.total_entries + 1), 0),
                        last_spin_date = CURRENT_TIMESTAMP
                    """,
                    user_id, total_spins
                )
        return True

    async def record_spin_prize(self, user_id: int, prize: str) -> None:
        """
        Persist prize history and optionally notify admin/log.
        """
        async with self._open_connection() as conn:
            await conn.execute(
                """
                INSERT INTO jobs_log (job_name, key, status)
                VALUES ('spin_prize', $1, $2)
                """,
                str(user_id), prize
            )

    async def sync_spins_for_user(self, user_id: int) -> None:
        """
        Recompute available_spins from bites and used spins and persist to spin_entries.
        Useful to run after bulk bites updates (orders/referrals).
        """
        async with self._open_connection() as conn:
            bites = await conn.fetchval("SELECT bites FROM leaderboards WHERE user_id=$1", user_id) or 0
            used = await conn.fetchval("SELECT total_entries FROM spin_entries WHERE user_id=$1", user_id) or 0
            total_spins = bites // 25
            available = max(0, total_spins - used)
            await conn.execute(
                """
                INSERT INTO spin_entries (user_id, total_entries, available_spins, last_spin_date)
                VALUES ($1, $2, $3, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id) DO UPDATE SET available_spins = $3
                """,
                user_id, used, available
            )


    async def get_genna_progress(self, user_id: int) -> dict:
        import datetime
        async with self._open_connection() as conn:
            today = datetime.date.today()
            monday = today - datetime.timedelta(days=today.weekday())
            sunday = monday + datetime.timedelta(days=6)

            # Bites earned this week
            bites_this_week = await conn.fetchval(
                """
                SELECT SUM(bites)
                FROM leaderboards
                WHERE user_id=$1 AND last_updated >= $2
                """,
                user_id, monday
            ) or 0

            # Current rank
            rank_position = await conn.fetchval(
                """
                SELECT r FROM (
                    SELECT user_id, RANK() OVER (ORDER BY bites DESC) AS r
                    FROM leaderboards
                ) t WHERE user_id=$1
                """,
                user_id
            ) or None

            # Highest bites in leaderboard
            highest_bites = await conn.fetchval(
                "SELECT MAX(bites) FROM leaderboards"
            ) or 0

            # If user is #1, no next threshold
            if bites_this_week >= highest_bites and rank_position == 1:
                next_threshold = None
                remaining = None
                message = "🏆 You are #1 on the leaderboard! Keep defending your spot."
            else:
                # Otherwise compute next threshold normally
                next_threshold = ((bites_this_week // 10) + 1) * 10
                remaining = max(0, next_threshold - bites_this_week)
                message = f"🎯 Next rank in: {remaining} Bites"

            days_left = (sunday - today).days

        return {
            "bites_this_week": bites_this_week,
            "rank_position": rank_position,
            "highest_bites": highest_bites,
            "next_threshold": next_threshold,
            "remaining": remaining,
            "days_left": days_left,
            "message": message
        }
        
    async def get_user_rank_and_progress(self, user_id: int):
        async with self._open_connection() as conn:
            rank = await conn.fetchval(
                """
                SELECT r FROM (
                    SELECT user_id, RANK() OVER (ORDER BY bites DESC) AS r
                    FROM leaderboards
                ) t WHERE user_id=$1
                """,
                user_id
            ) or None
            # progress to next threshold
            bites = await conn.fetchval("SELECT bites FROM leaderboards WHERE user_id=$1", user_id) or 0
            next_threshold = ((bites // 10) + 1) * 10
            progress_pct = min(100.0, (bites / next_threshold) * 100)
        return rank, progress_pct



    async def count_new_users(self, date: str) -> int:
        """Count users who signed up on a given date."""
        async with self._open_connection() as conn:
            return await conn.fetchval(
                "SELECT COUNT(*) FROM users WHERE created_at >= $1 AND created_at < $1 + interval '1 day'",
                date
            )
            

    async def summarize_orders_day(self, date: str) -> Dict[str, Any]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT COUNT(*) AS total,
                    SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) AS delivered,
                    SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) AS cancelled,
                    COALESCE(SUM(CASE WHEN status='delivered' THEN food_subtotal ELSE 0 END),0) AS food_revenue,
                    COALESCE(SUM(CASE WHEN status='delivered' THEN delivery_fee ELSE 0 END),0) AS delivery_fees
                FROM orders
                WHERE created_at::DATE = $1
                """,
                date
            )
            delivered = row["delivered"] or 0
            cancelled = row["cancelled"] or 0
            denom = delivered + cancelled
            reliability_pct = 0 if denom == 0 else round(100.0 * delivered / denom)
            return {
                "total": int(row["total"]),
                "delivered": int(delivered),
                "cancelled": int(cancelled),
                "food_revenue": float(row["food_revenue"]),   # delivered only
                "delivery_fees": float(row["delivery_fees"]), # delivered only
                "reliability_pct": reliability_pct,
            }
                
            # List cancelled orders with meal/vendor names
            
    async def get_active_orders_for_dg(self, dg_id: int) -> List[Dict]:
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                "SELECT * FROM orders WHERE delivery_guy_id = $1 AND status IN ('assigned','in_progress','ready')",
                dg_id
            )
            return [self._row_to_dict(r) for r in rows]

    async def list_cancelled_orders_day(self, date: str):
        async with self._pool.acquire() as conn:
            return await conn.fetch(
                """
                SELECT o.user_id,
                    v.name AS vendor_name,
                    item->>'name' AS meal_name
                FROM orders o
                JOIN vendors v ON o.vendor_id = v.id
                CROSS JOIN LATERAL jsonb_array_elements(o.items_json::jsonb) AS item
                WHERE o.status='cancelled' AND o.created_at::DATE=$1
                """,
                date
            )
            
        # in db.py
    async def get_platform_total(self, day: datetime.date) -> Optional[float]:
        """Returns the total platform profit for all DGs on a given day."""
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT SUM(earnings) AS total_profit
                FROM daily_stats
                WHERE date = $1
                """,
                day
            )
            return row["total_profit"] if row and row["total_profit"] is not None else 0.0


    # Top delivered meal
    async def top_meal_day(self, date: str):
        async with self._pool.acquire() as conn:
            row = await conn.fetchrow(
                """
                SELECT item->>'name' AS meal_name, COUNT(*) AS cnt
                FROM orders o
                CROSS JOIN LATERAL jsonb_array_elements(o.items_json::jsonb) AS item
                WHERE o.status='delivered' AND o.created_at::DATE=$1
                GROUP BY meal_name
                ORDER BY cnt DESC
                LIMIT 1
                """,
                date
            )
            return (row["meal_name"], row["cnt"]) if row else ("None", 0)
        
    
    async def get_user_stats(self, telegram_id: int) -> Optional[Dict[str, Any]]:
        """
        Fetch user stats including username, xp, level, and order_count.
        """
        async with self._open_connection() as conn:
            user = await conn.fetchrow(
                "SELECT id, telegram_id, first_name, phone, campus, xp, level "
                "FROM users WHERE telegram_id=$1",
                telegram_id
            )
            if not user:
                return None

            # Count orders for this user
            order_count = await conn.fetchval(
                "SELECT COUNT(*) FROM orders WHERE user_id=$1",
                user["id"]
            )

            return {
                "id": user["id"],
                "telegram_id": user["telegram_id"],
                "first_name": user["first_name"],
                "phone": user["phone"],
                "campus": user["campus"],
                "xp": user["xp"],
                "level": user["level"],
                "order_count": order_count,
            }

    
    async def count_today_orders_for_dg(self, dg_id: int) -> Tuple[int, int]:
        """
        Returns (active_count, in_progress_count) for today.
        active_count = pending/assigned/preparing/ready
        in_progress_count = in_progress/on_the_way
        """
        today = date.today()
        sql = """
            SELECT
                SUM(CASE WHEN status IN ('pending','assigned','preparing','ready') THEN 1 ELSE 0 END) AS active_count,
                SUM(CASE WHEN status IN ('in_progress','on_the_way') THEN 1 ELSE 0 END) AS in_progress_count
            FROM orders
            WHERE delivery_guy_id = $1
            AND DATE(created_at) = $2
        """
        row = await self._pool.fetchrow(sql, dg_id, today)
        return int(row["active_count"] or 0), int(row["in_progress_count"] or 0)


    # Top vendor delivered
    async def top_vendor_delivered_day(self, date: str):
        async with self._pool.acquire() as conn:
            row = await conn.fetchrow(
                """
                SELECT v.name AS vendor_name, COUNT(*) AS cnt
                FROM orders o
                JOIN vendors v ON o.vendor_id = v.id
                WHERE o.status='delivered' AND o.created_at::DATE=$1
                GROUP BY vendor_name
                ORDER BY cnt DESC
                LIMIT 1
                """,
                date
            )
            return (row["vendor_name"], row["cnt"]) if row else ("None", 0)

    # Top vendor cancelled
    async def top_vendor_cancelled_day(self, date: str):
        async with self._pool.acquire() as conn:
            row = await conn.fetchrow(
                """
                SELECT v.name AS vendor_name, COUNT(*) AS cnt
                FROM orders o
                JOIN vendors v ON o.vendor_id = v.id
                WHERE o.status='cancelled' AND o.created_at::DATE=$1
                GROUP BY vendor_name
                ORDER BY cnt DESC
                LIMIT 1
                """,
                date
            )
            return (row["vendor_name"], row["cnt"]) if row else ("None", 0)

   # db/tickets.py
    async def save_ticket(self, ticket_id, user_id, text, status, original_msg_id):
        async with self._pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO tickets (ticket_id, user_id, text, status, original_msg_id)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (ticket_id) DO NOTHING
            """, ticket_id, user_id, text, status, original_msg_id)

    async def get_ticket(self, ticket_id):
        async with self._pool.acquire() as conn:
            return await conn.fetchrow(
                "SELECT ticket_id, user_id, text, status, original_msg_id FROM tickets WHERE ticket_id=$1",
                ticket_id
            )

    async def list_open_tickets(self):
        async with self._pool.acquire() as conn:
            return await conn.fetch(
                "SELECT ticket_id, status FROM tickets WHERE status='open' ORDER BY created_at DESC"
            )

    async def close_ticket(self, ticket_id):
        async with self._pool.acquire() as conn:
            await conn.execute(
                "UPDATE tickets SET status='closed', closed_at=CURRENT_TIMESTAMP WHERE ticket_id=$1",
                ticket_id
            )
            
    
    async def list_closed_tickets(self):
        async with self._pool.acquire() as conn:
            return await conn.fetch(
                "SELECT ticket_id, status, user_id, text, closed_at "
                "FROM tickets WHERE status='closed' ORDER BY closed_at DESC"
            )
            
    
    
            
            
    
    
    async def summarize_vendors_day(self, date: str) -> Dict[str, Any]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT COUNT(*) AS active,
                    AVG(rating_avg) AS avg_rating
                FROM vendors
                WHERE status = 'active'
                """,
            )
            return {
                "active": int(row["active"] or 0),
                "avg_rating": float(row["avg_rating"] or 0.0),
            }
            
    
    async def summarize_delivery_day(self, date: str) -> Dict[str, Any]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT COUNT(*) AS active,
                    SUM(total_deliveries) AS deliveries,
                    AVG(
                        CASE WHEN total_requests > 0
                                THEN accepted_requests::float / total_requests
                                ELSE NULL END
                    ) AS acceptance_rate
                FROM delivery_guys
                WHERE active = TRUE
                """,
            )
            return {
                "active": int(row["active"] or 0),
                "deliveries": int(row["deliveries"] or 0),
                "acceptance_rate": float(row["acceptance_rate"] or 0.0) * 100,
            }
            
    
    async def top_campus_day(self, date: str) -> Tuple[str, int]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT campus, COUNT(*) AS orders
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.created_at::DATE = $1
                GROUP BY campus
                ORDER BY orders DESC
                LIMIT 1
                """,
                date
            )
            if row:
                return row["campus"], int(row["orders"])
            return "N/A", 0



    async def get_orders_for_vendor(
        self,
        vendor_id: int,
        *,
        date: Optional[str] = None,
        status_filter: Optional[List[str]] = None,
        limit: Optional[int] = None,
        offset: int = 0
    ) -> List[Dict[str, Any]]:
        """
        List orders for a vendor, optionally filtered by date (YYYY-MM-DD),
        status list, and paginated (limit/offset).
        """
        where = ["vendor_id = $1"]
        params: List[Any] = [vendor_id]
        param_counter = 2

        from datetime import datetime, date as date_cls

        if date:
            where.append(f"created_at >= ${param_counter} AND created_at < ${param_counter + 1}")

            # normalize: accept either str "YYYY-MM-DD" or datetime.date
            if isinstance(date, str):
                base_date = datetime.strptime(date, "%Y-%m-%d").date()
            elif isinstance(date, datetime):
                base_date = date.date()
            elif isinstance(date, date_cls):
                base_date = date
            else:
                raise TypeError("date must be str, datetime, or date")

            start_date = datetime.combine(base_date, datetime.min.time())
            end_date   = datetime.combine(base_date, datetime.max.time())

            params.extend([start_date, end_date])
            param_counter += 2

        if status_filter:
            where.append(f"status = ANY(${param_counter})")
            params.append(status_filter)
            param_counter += 1

        sql = f"""
            SELECT * FROM orders
            WHERE {' AND '.join(where)}
            ORDER BY created_at DESC
        """
        if limit is not None:
            sql += f" LIMIT ${param_counter} OFFSET ${param_counter + 1}"
            params.extend([limit, offset])

        rows = await self._pool.fetch(sql, *params)
        return [self._row_to_dict(r) for r in rows]


    async def count_orders_for_vendor(
        self,
        vendor_id: int,
        *,
        date: Optional[str] = None,
        status_filter: Optional[List[str]] = None
    ) -> int:
        """Count orders for pagination and summaries."""
        where = ["vendor_id = $1"]
        params: List[Any] = [vendor_id]
        param_counter = 2

        from datetime import datetime

        if date:
            where.append(f"created_at >= ${param_counter} AND created_at < ${param_counter + 1}")
            # if date is already a str like "2025-12-22", parse it
            if isinstance(date, str):
                base_date = datetime.strptime(date, "%Y-%m-%d").date()
            else:
                base_date = date  # already a datetime.date

            start_date = datetime.combine(base_date, datetime.min.time())
            end_date = datetime.combine(base_date, datetime.max.time())
            params.extend([start_date, end_date])
            param_counter += 2
        if status_filter:
            where.append(f"status = ANY(${param_counter})")
            params.append(status_filter)
            param_counter += 1

        sql = f"SELECT COUNT(*) FROM orders WHERE {' AND '.join(where)}"
        count = await self._pool.fetchval(sql, *params)
        return int(count)

    async def summarize_vendor_day(self, vendor_id: int, date: Optional[str] = None) -> Dict[str, Any]:
        """
        Returns daily summary: counts, food revenue, delivery fees, ratings snapshot, reliability.
        date format: YYYY-MM-DD (defaults to today if None).
        """
        date = date or datetime.date.today().strftime("%Y-%m-%d")
        async with self._open_connection() as conn:

            # Orders summary (delivered vs cancelled vs prepared)
            s = await conn.fetchrow(
                """
                SELECT
                SUM(CASE WHEN status = 'delivered' THEN 1 ELSE 0 END) AS delivered_count,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled_count,
                COALESCE(SUM(food_subtotal), 0.0) AS food_revenue,
                COALESCE(SUM(delivery_fee), 0.0) AS delivery_fees
                FROM orders
                WHERE vendor_id = $1 AND created_at::DATE = $2
                """,
                vendor_id, date
            )
            delivered = int(s["delivered_count"] if s else 0)
            cancelled = int(s["cancelled_count"] if s else 0)
            food_rev = float(s["food_revenue"] if s else 0.0)
            delivery_fees = float(s["delivery_fees"] if s else 0.0)

            # Ratings snapshot (from cached vendor fields)
            v = await conn.fetchrow("SELECT rating_avg, rating_count FROM vendors WHERE id = $1", vendor_id)
            rating_avg = float(v["rating_avg"] if v else 0.0)
            rating_count = int(v["rating_count"] if v else 0)

            # Reliability calculation
            r = await conn.fetchrow(
                """
                SELECT
                SUM(CASE WHEN status IN ('preparing','ready','in_progress','delivered') THEN 1 ELSE 0 END) AS progressed,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled
                FROM orders
                WHERE vendor_id = $1 AND created_at::DATE = $2
                """,
                vendor_id, date
            )
            progressed = int(r["progressed"] or 0)
            cancelled2 = int(r["cancelled"] or 0)
            denom = progressed + cancelled2
            reliability = 0 if denom == 0 else round(100.0 * progressed / denom)

        return {
            "date": date,
            "delivered": delivered,
            "cancelled": cancelled,
            "food_revenue": food_rev,
            "delivery_fees": delivery_fees,
            "total_payout": food_rev + delivery_fees,
            "rating_avg": rating_avg,
            "rating_count": rating_count,
            "reliability_pct": reliability,
        }

    async def summarize_vendor_week(self, vendor_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Weekly summary across a date range (defaults to current ISO week).
        Returns totals and per-day breakdown.
        """
        if not start_date or not end_date:
            today = datetime.date.today()
            start = today - datetime.timedelta(days=today.weekday())  # Monday
            end = start + datetime.timedelta(days=6)  # Sunday
            start_date = start.strftime("%Y-%m-%d")
            end_date = end.strftime("%Y-%m-%d")

        async with self._open_connection() as conn:

            # Totals
            t = await conn.fetchrow(
                """
                SELECT
                SUM(CASE WHEN status = 'delivered' THEN 1 ELSE 0 END) AS delivered_count,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled_count,
                COALESCE(SUM(food_subtotal), 0.0) AS food_revenue,
                COALESCE(SUM(delivery_fee), 0.0) AS delivery_fees
                FROM orders
                WHERE vendor_id = $1 AND created_at::DATE BETWEEN $2 AND $3
                """,
                vendor_id, start_date, end_date
            )
            delivered = int(t["delivered_count"] or 0)
            cancelled = int(t["cancelled_count"] or 0)
            food_rev = float(t["food_revenue"] or 0.0)
            delivery_fees = float(t["delivery_fee"] or 0.0)

            # Per-day breakdown
            rows = await conn.fetch(
                """
                SELECT created_at::DATE AS d,
                    SUM(CASE WHEN status = 'delivered' THEN 1 ELSE 0 END) AS delivered_count,
                    SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled_count,
                    COALESCE(SUM(food_subtotal), 0.0) AS food_revenue,
                    COALESCE(SUM(delivery_fee), 0.0) AS delivery_fees
                FROM orders
                WHERE vendor_id = $1 AND created_at::DATE BETWEEN $2 AND $3
                GROUP BY created_at::DATE
                ORDER BY d ASC
                """,
                vendor_id, start_date, end_date
            )
            days = [{
                "date": str(r["d"]), # Convert date object to string
                "delivered": int(r["delivered_count"] or 0),
                "cancelled": int(r["cancelled_count"] or 0),
                "food_revenue": float(r["food_revenue"] or 0.0),
                "delivery_fees": float(r["delivery_fees"] or 0.0),
                "total_payout": float(r["food_revenue"] or 0.0) + float(r["delivery_fees"] or 0.0),
            } for r in rows]

        return {
            "start_date": start_date,
            "end_date": end_date,
            "delivered": delivered,
            "cancelled": cancelled,
            "food_revenue": food_rev,
            "delivery_fees": delivery_fees,
            "total_payout": food_rev + delivery_fees,
            "days": days
        }


    async def get_vendor_daily_orders_page(self, vendor_id: int, date: Optional[str], page: int, page_size: int = 10) -> Dict[str, Any]:
        """
        Returns {orders, total, page, pages} for the given day.
        """
        date = date or datetime.date.today().strftime("%Y-%m-%d")
        total = await self.count_orders_for_vendor(vendor_id, date=date)
        pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, pages))
        offset = (page - 1) * page_size
        orders = await self.get_orders_for_vendor(vendor_id, date=date, limit=page_size, offset=offset)
        return {"orders": orders, "total": total, "page": page, "pages": pages}


    async def get_vendor_weekly_orders_page(self, vendor_id: int, start_date: Optional[str], end_date: Optional[str], page: int, page_size: int = 10) -> Dict[str, Any]:
        """
        Paginates across the week range.
        NOTE: This implementation does not filter by date in get_orders_for_vendor,
        but the count is for the range. Assuming you want all orders for the vendor.
        If you want to filter get_orders_for_vendor, you'd need to modify it to accept the range.
        For now, I'm fetching the count for the period and paginating all orders (as the original get_orders_for_vendor was called).
        """
        
        async with self._open_connection() as conn:
            where = ["vendor_id = $1", "created_at::DATE BETWEEN $2 AND $3"]
            params = [vendor_id, start_date, end_date]
            total = await conn.fetchval(f"SELECT COUNT(*) FROM orders WHERE {' AND '.join(where)}", *params)
            total = int(total) if total is not None else 0


        pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, pages))
        offset = (page - 1) * page_size
        
        # NOTE: The original code called get_orders_for_vendor without the date filter here,
        # which would fetch *all* vendor orders, paginated.
        # I'm retaining the original *behavior* but the count is for the range.
        # If you need orders *only* for the range, modify the call below:
        orders = await self.get_orders_for_vendor(vendor_id, limit=page_size, offset=offset)
        
        return {"orders": orders, "total": total, "page": page, "pages": pages, "start_date": start_date, "end_date": end_date}

    async def calc_vendor_reliability_for_day(self, vendor_id: int, date: Optional[str] = None) -> float:
        """
        Returns reliability percentage for a given day:
        progressed (accepted/in_progress/delivered) vs cancelled.
        """
        import datetime
        date = datetime.date.today()
        async with self._open_connection() as conn:
            r = await conn.fetchrow(
                """
                SELECT
                SUM(CASE WHEN status IN ('preparing','ready','in_progress','delivered') THEN 1 ELSE 0 END) AS progressed,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled
                FROM orders
                WHERE vendor_id = $1 AND created_at::DATE = $2
                """,
                vendor_id, date
            )
            progressed = int(r["progressed"] or 0)
            cancelled = int(r["cancelled"] or 0)

        denom = progressed + cancelled
        return 0.0 if denom == 0 else round(100.0 * progressed / denom, 2)
    
    async def get_delivery_guy_by_user_id(self, user_id: int):
        row = await self._pool.fetchrow(
            "SELECT id, user_id, telegram_id, name FROM delivery_guys WHERE user_id = $1 LIMIT 1",
            user_id
        )
        if row is None:
            raise ValueError(f"Delivery guy with user_id={user_id} not found in DB")
        return dict(row)

    async def get_delivery_guy_by_id(self, dg_id: int):
        row = await self._pool.fetchrow(
            "SELECT id, user_id, telegram_id, name FROM delivery_guys WHERE id = $1 LIMIT 1",
            dg_id
        )
        if row is None:
            raise ValueError(f"Delivery guy with id={dg_id} not found in DB")
        return dict(row)

    async def get_delivery_guy_telegram_id_by_id(self, dg_id: int) -> int:
        row = await self._pool.fetchrow(
            "SELECT telegram_id FROM delivery_guys WHERE id = $1 LIMIT 1",
            dg_id
        )
        if row is None or row["telegram_id"] is None:
            raise ValueError(f"telegram_id for delivery_guy_id={dg_id} not found")
        return int(row["telegram_id"])

    async def get_delivery_guy_telegram_id(self, user_id: int) -> int:
        row = await self._pool.fetchrow(
            "SELECT telegram_id FROM delivery_guys WHERE user_id = $1 LIMIT 1",
            user_id
        )
        if row is None or row["telegram_id"] is None:
            raise ValueError(f"telegram_id for user_id={user_id} not found")
        return int(row["telegram_id"])

    async def get_delivery_guy(self, delivery_guy_id: int) -> Optional[Dict[str, Any]]:
        row = await self._pool.fetchrow(
            "SELECT * FROM delivery_guys WHERE id = $1 LIMIT 1",
            delivery_guy_id
        )
        return self._row_to_dict(row) if row else None
    
    async def create_delivery_guy(
        self,
        telegram_id: int,
        name: str,
        campus: str,
        gender: str = None,
        phone: str = None
    ) -> int:
        async with self._pool.acquire() as conn:
            async with conn.transaction():
                # 1. Create or update user record with role=delivery_guy
                user_id = await conn.fetchval(
                    """
                    INSERT INTO users (telegram_id, role, first_name, phone, campus)
                    VALUES ($1, 'delivery_guy', $2, $3, $4)
                    ON CONFLICT (telegram_id) DO UPDATE
                    SET role = 'delivery_guy',
                        first_name = EXCLUDED.first_name,
                        phone = EXCLUDED.phone,
                        campus = EXCLUDED.campus
                    RETURNING id
                    """,
                    telegram_id, name, phone, campus
                )

                # 2. Create or update delivery_guy record linked to that user
                dg_id = await conn.fetchval(
                    """
                    INSERT INTO delivery_guys (user_id, telegram_id, name, campus, gender, phone, active, total_deliveries)
                    VALUES ($1, $2, $3, $4, $5, $6, TRUE, 0)
                    ON CONFLICT (user_id) DO UPDATE
                    SET telegram_id = EXCLUDED.telegram_id,
                        name = EXCLUDED.name,
                        campus = EXCLUDED.campus,
                        gender = EXCLUDED.gender,
                        phone = EXCLUDED.phone,
                        active = FALSE
                    RETURNING id
                    """,
                    user_id, telegram_id, name, campus, gender, phone
                )

        return int(dg_id) if dg_id else 0


    async def get_delivery_guy_by_user(self, telegram_id: int):
        # Remove pool recycle; rely on asyncpg’s internal statement cache
        return await self._pool.fetchrow(
            "SELECT * FROM delivery_guys WHERE telegram_id = $1 LIMIT 1",
            telegram_id
        )
        
    async def get_delivery_guy_by_user_onboard(self, telegram_id: int):
        return await self._pool.fetchrow(
            "SELECT * FROM delivery_guys WHERE telegram_id = $1 LIMIT 1",
            telegram_id
        )




    async def get_daily_stats_for_dg(self, dg_id: int, date: str) -> Dict[str, Any]:
        """ 
        Returns stats for a delivery guy on a given date.
        Includes delivered orders only.
        """
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT 
                    COALESCE(SUM(delivery_fee), 0) AS earnings,
                    COUNT(*) AS deliveries
                FROM orders
                WHERE delivery_guy_id = $1
                AND DATE(updated_at) = $2
                AND status = 'delivered'
                """,
                dg_id,
                date
            )


        if not row:
            return {"earnings": 0, "deliveries": 0}
        
        # XP and coins can be derived here
        xp = row["deliveries"] * 10
        coins = row["earnings"] * 0.05
        return {
            "earnings": row["earnings"],
            "deliveries": row["deliveries"],
            "xp": xp,
            "coins": coins
        }
    
    async def get_top_drivers(self, date: str, limit: int = 3) -> list[dict]:
        """
        Returns the top drivers for a given date, ranked by deliveries (then earnings).
        Uses get_daily_stats_for_dg to compute stats directly from orders.
        """
        async with self._open_connection() as conn:
            # Fetch all delivery guys
            dgs = await conn.fetch("SELECT id, name FROM delivery_guys")

        results = []
        for dg in dgs:
            stats = await self.get_daily_stats_for_dg(dg["id"], date)
            results.append({
                "id": dg["id"],
                "name": dg["name"] or f"DG #{dg['id']}",
                "deliveries": stats["deliveries"],
                "earnings": stats["earnings"],
                "xp": stats["xp"],
                "coins": stats["coins"],
            })

        # Sort by deliveries first, then earnings
        results.sort(key=lambda r: (r["deliveries"], r["earnings"]), reverse=True)

        return results[:limit]

    async def get_weekly_earnings_for_dg(self, dg_id: int, week_start: str, week_end: str) -> List[Dict[str, Any]]:
        """
        Returns day-by-day breakdown for the week.
        Each entry: {date, deliveries, earnings, xp, coins}
        """
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                """
                SELECT 
                    DATE(updated_at) AS day,
                    COALESCE(SUM(delivery_fee), 0) AS earnings,
                    COUNT(*) AS deliveries
                FROM orders
                WHERE delivery_guy_id = $1
                AND status = 'delivered'
                AND DATE(updated_at) BETWEEN $2 AND $3
                GROUP BY day
                ORDER BY day ASC
                """,
                dg_id, week_start, week_end
            )
        
        stats = []
        for r in rows:
            xp = r["deliveries"] * 10
            coins = r["earnings"] * 0.05
            stats.append({
                "date": r["day"],
                "deliveries": r["deliveries"],
                "earnings": r["earnings"],
                "xp": xp,
                "coins": coins
            })
        return stats
    async def get_weekly_totals_for_dg(self, dg_id: int, week_start: str, week_end: str) -> Dict[str, Any]:
        """
        Returns total deliveries, earnings, xp, coins for the week.
        """
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
    """
    SELECT 
        COALESCE(SUM(delivery_fee), 0) AS earnings,
        COUNT(*) AS deliveries
    FROM orders
    WHERE delivery_guy_id = $1
      AND status = 'delivered'
      AND DATE(updated_at) BETWEEN $2 AND $3
    """,
    dg_id, week_start, week_end
)

        if not row:
            return {"earnings": 0, "deliveries": 0, "xp": 0, "coins": 0}
        
        xp = row["deliveries"] * 10
        coins = row["earnings"] * 0.05
        return {
            "earnings": row["earnings"],
            "deliveries": row["deliveries"],
            "xp": xp,
            "coins": coins
        }

    async def update_delivery_guy_coords(self, dg_id: int, lat: float, lon: float):
        """
        Update the last known coordinates of a delivery guy.
        """
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE delivery_guys SET last_lat = $1, last_lon = $2 WHERE id = $3",
                lat, lon, dg_id
            )
    
    # --- New Status/Timestamp Methods ---
    async def set_delivery_guy_online(self, dg_id: int) -> None:
        """Sets active=1 and updates last_online_at."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE delivery_guys SET active = TRUE, last_online_at = CURRENT_TIMESTAMP WHERE id = $1",
                dg_id
            )
            
    
    # in your db layer
    async def list_all_users(self) -> List[Dict[str, Any]]:
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                "SELECT id, telegram_id FROM users WHERE status = 'active' AND role = 'student'"
            )
            return [dict(r) for r in rows]
        
    
    # async def set_user_opt_out(self, telegram_id: int, opt_out: bool):
    #     async with self._open_connection() as conn:
    #         await conn.execute(
    #             "UPDATE users SET opt_out_reminders = $1 WHERE telegram_id = $2",
    #             opt_out, telegram_id
    #         )





    # in your db layer
    async def list_active_students(self) -> List[Dict[str, Any]]:
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                """
                SELECT id, telegram_id, first_name
                FROM users
                WHERE status = 'active'
                AND role = 'student'
         
                """
            )
            return [dict(r) for r in rows]
        
    
    async def list_active_studentss(self) -> List[Dict[str, Any]]:
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                """
                SELECT id, telegram_id, first_name
                FROM users
                WHERE status = 'active'
                  AND role = 'student'
                """
            )
            return [dict(r) for r in rows]
    

    async def set_delivery_guy_offline(self, dg_id: int) -> None:
        """Sets active=0 and updates last_offline_at."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE delivery_guys SET active = FALSE, last_offline_at = CURRENT_TIMESTAMP WHERE id = $1",
                dg_id
            )

    async def block_delivery_guy(self, dg_id: int) -> None:
        """Sets active=FALSE and blocked=TRUE."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE delivery_guys SET active = FALSE, blocked = TRUE WHERE id = $1",
                dg_id
            )

    # -------------------- Orders --------------------
    async def create_order(
        self,
        user_id: int,
        vendor_id: int,
        pickup: str,
        dropoff: str,
        items_json: str,
        food_subtotal: float,
        delivery_fee: float,
        status: str,
        payment_method: str,
        payment_status: str,
        receipt_id: Optional[int],
        breakdown_json: str,
        notes: Optional[str] = None,
        delivery_guy_id: Optional[int] = None,
    ) -> int:
        from datetime import datetime, timedelta, timezone
        now = datetime.utcnow()  # naive UTC
        expires_at = now + timedelta(minutes=45)
        async with self._open_connection() as conn:
            
            order_id = await conn.fetchval(
                """
                INSERT INTO orders
                (user_id, delivery_guy_id, vendor_id, pickup, dropoff, items_json,
                food_subtotal, delivery_fee, status, payment_method, payment_status,
                receipt_id, breakdown_json, notes, created_at, updated_at, expires_at)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$15,$16)
                RETURNING id
                """,
                user_id, delivery_guy_id, vendor_id, pickup, dropoff, items_json,
                food_subtotal, delivery_fee, status, payment_method, payment_status,
                receipt_id, breakdown_json, notes, now, expires_at
            )

            return int(order_id) if order_id is not None else 0

    from datetime import datetime
    from aiogram import Bot
    async def get_internal_user_id(self, telegram_id: int) -> Optional[int]:
        row = await self._pool.fetchrow(
            "SELECT id FROM users WHERE telegram_id = $1",
            telegram_id
        )
        return row["id"] if row else None
    
    
    async def get_user_by_id(self, internal_user_id: int) -> Optional[Dict[str, Any]]:
        """Return the users row by internal DB id."""
        row = await self._pool.fetchrow(
            "SELECT * FROM users WHERE id = $1 LIMIT 1",
            internal_user_id
        )
        return dict(row) if row else None


    async def check_thresholds_and_notify(
        self,
        bot: Bot,
        dg_id: int,
        admin_group_id: int,
        max_skips: int = 3
    ):
        """Check today's skip count and notify DG/admin if threshold exceeded."""
        today = datetime.now().strftime("%Y-%m-%d")

        async with self._open_connection() as conn:
            stats = await conn.fetchrow(
                "SELECT skipped FROM daily_stats WHERE dg_id=$1 AND date=$2",
                dg_id, today
            )
            skips_today = int(stats["skipped"]) if stats else 0

        # Fetch DG info with telegram_id
        try:
            dg_info = await self.get_delivery_guy_by_id(dg_id)
        except ValueError:
            print(f"[check_thresholds_and_notify] DG {dg_id} not found in DB")
            return

        tg_id = dg_info.get("telegram_id")
        name = dg_info.get("name")

        if not tg_id:
            print(f"[check_thresholds_and_notify] DG {dg_id} has no telegram_id stored")
            return

        # DG feedback
        try:
            if skips_today == 0:
                await bot.send_message(
                    tg_id,
                    "✅ Great job! No skips today — keep it up 🚀"
                )
            elif skips_today < max_skips:
                await bot.send_message(
                    tg_id,
                    f"⚠️ You’ve skipped {skips_today} orders today.\n"
                    f"Stay reliable to keep receiving orders."
                )
            else:
                await bot.send_message(
                    tg_id,
                    f"🚨 You’ve reached the skip limit ({max_skips}).\n"
                    f"Further skips may affect your reliability."
                )
        except Exception as e:
            print(f"Failed to notify DG {dg_id} ({name}) about skips: {e}")

        # Admin alert if threshold exceeded
        if skips_today >= max_skips:
            admin_message = (
                f"🚨 **Reliability Alert!**\n"
                f"Delivery Partner **{name}** (ID: {dg_id}) has reached {skips_today} skips today.\n"
                f"**ACTION REQUIRED**: Review their performance and block if necessary."
            )
            try:
                await bot.send_message(admin_group_id, admin_message, parse_mode="Markdown")
            except Exception as e:
                print(f"Failed to notify admin group {admin_group_id}: {e}")
                
    async def record_daily_stat_delivery(self, dg_id: int, date_str: str, earnings: float, total_xp: int = 10, total_coins: float = 0.0) -> None:
        """Updates daily_stats and dg gamification stats upon a successful delivery."""
        
        async with self._open_connection() as conn:
            # 1. Update daily_stats table
            await conn.execute(
                """
                INSERT INTO daily_stats (dg_id, date, deliveries, earnings)
                VALUES ($1, $2, 1, $3)
                ON CONFLICT(dg_id, date) DO UPDATE SET
                deliveries = daily_stats.deliveries + 1,
                earnings = daily_stats.earnings + $3,
                updated_at = CURRENT_TIMESTAMP
                """,
                dg_id, date_str, earnings
            )
            
            # 2. Update delivery_guys gamification stats and total deliveries
            await conn.execute(
                """
                UPDATE delivery_guys SET
                total_deliveries = total_deliveries + 1,
                xp = xp + $1,
                coins = coins + $2
                WHERE id = $3
                """,
                total_xp, total_coins, dg_id
            )
    
    async def increment_total_requests(self, dg_id: int) -> None:
            """Increment total_requests whenever a new order offer is sent to a DG."""
            async with self._open_connection() as conn:
                await conn.execute(
                    """
                    UPDATE delivery_guys
                    SET total_requests = total_requests + 1
                    WHERE id = $1
                    """,
                    dg_id
                )

    async def increment_accepted_requests(self, dg_id: int) -> None:
        """Increment accepted_requests when a DG accepts an order offer."""
        async with self._open_connection() as conn:
            await conn.execute(
                """
                UPDATE delivery_guys
                SET accepted_requests = accepted_requests + 1
                WHERE id = $1
                """,
                dg_id
            )

    
    async def increment_skip(self, dg_id: int) -> None:
        """Increment both lifetime skipped_requests and today's skipped count."""
        today = datetime.now().strftime("%Y-%m-%d")
        async with self._open_connection() as conn:
            # 1. Update lifetime stats
            await conn.execute(
                """
                UPDATE delivery_guys
                SET skipped_requests = skipped_requests + 1,
                    last_skip_at = CURRENT_TIMESTAMP
                WHERE id = $1
                """,
                dg_id
            )

            # 2. Upsert into daily_stats
            await conn.execute(
                """
                INSERT INTO daily_stats (dg_id, date, skipped, updated_at)
                VALUES ($1, $2, 1, CURRENT_TIMESTAMP)
                ON CONFLICT (dg_id, date)
                DO UPDATE SET skipped = daily_stats.skipped + 1,
                            updated_at = CURRENT_TIMESTAMP
                """,
                dg_id, today
            )


    async def increment_total_deliveries(self, dg_id: int) -> None:
        import datetime
        """Increment total_deliveries when a DG successfully delivers an order and update daily_stats."""
        today_str = datetime.date.today().strftime("%Y-%m-%d")

        # 1. Update delivery_guys cumulative counter
        async with self._open_connection() as conn:
            await conn.execute(
                """
                UPDATE delivery_guys
                SET total_deliveries = total_deliveries + 1
                WHERE id = $1
                """,
                dg_id
            )
    
    

        # 2. Update daily_stats table (UPSERT)
        async with self._open_connection() as conn:
            await conn.execute(
                """
                INSERT INTO daily_stats (dg_id, date, deliveries)
                VALUES ($1, $2, 1)
                ON CONFLICT(dg_id, date) DO UPDATE SET
                    deliveries = daily_stats.deliveries + 1,
                    updated_at = CURRENT_TIMESTAMP
                """,
                dg_id, today_str
            )


    async def get_user_campus_by_order(self, order_id: int) -> Optional[str]:
        """
        Fetch the campus of the user who placed a given order
        and return it as text with the corresponding emoji.
        
        Args:
            order_id (int): The ID of the order.

        Returns:
            Optional[str]: Campus name with emoji, or None if not found.
        """
        # Campus → emoji mapping
        campus_emojis = {
            "4kilo": "🏛",
            "5kilo": "📚",
            "6kilo": "🎓",
            "FBE": "💹"
        }

        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                """
                SELECT u.campus
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.id = $1
                """,
                order_id
            )

            if not row:
                return None

            campus = row["campus"]
            emoji = campus_emojis.get(campus, "")
            return f"{emoji} {campus}"

    async def reset_daily_skip_count(self, dg_id: int) -> None:
        """Resets the DG's `skipped_requests` counter."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE delivery_guys SET skipped_requests = 0 WHERE id = $1",
                dg_id
            )

    async def get_daily_stats(self, dg_id: int, date_str: str) -> Optional[Dict[str, Any]]:
        """Retrieves daily stats for a DG on a specific date."""
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM daily_stats WHERE dg_id = $1 AND date = $2",
                dg_id, date_str
            )
            return self._row_to_dict(row) if row else None
        
    async def get_leaderboard(self, limit: int = 100) -> List[Dict[str, Any]]:
            """
            Return top users by XP for leaderboard display.
            """
            async with self._open_connection() as conn:
                rows = await conn.fetch(
                    """
                    SELECT id, telegram_id, first_name, xp, coins, campus, phone,
                        updated_at AS last_active
                    FROM users
                    ORDER BY xp DESC
                    LIMIT $1
                    """,
                    limit
                )
                return [self._row_to_dict(r) for r in rows]
    async def get_student_chat_id(self, order: Dict[str, Any]) -> Optional[int]:
        """
        Resolve the student's Telegram chat_id from an order record.
        - order["user_id"] is the internal DB id of the user.
        - This method fetches the user row and returns user["telegram_id"].
        """
        user_id = order.get("user_id")
        if not user_id:
            return None

        user = await self.get_user_by_id(user_id)
        return int(user["telegram_id"]) if user and user.get("telegram_id") is not None else None

    async def get_stats_for_period(self, dg_id: int, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """Retrieves stats for a delivery guy over a period."""
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                "SELECT * FROM daily_stats WHERE dg_id = $1 AND date BETWEEN $2 AND $3 ORDER BY date DESC",
                dg_id, start_date, end_date
            )
            return [self._row_to_dict(r) for r in rows]

    # --- Vendor, Location, & Other Methods (Adapted) ---

    
    async def list_vendors(self) -> List[Dict[str, Any]]:
        """Return all active vendors ordered by name."""
        async with self._open_connection() as conn:
            rows = await conn.fetch(
                "SELECT * FROM vendors WHERE status = 'active' ORDER BY name ASC"
            )
            return [self._row_to_dict(r) for r in rows]


    
    


# -------------------- Vendors --------------------

    async def get_vendor(self, vendor_id: int) -> Optional[Dict[str, Any]]:
        """Fetch a vendor by internal id."""
        async with self._open_connection() as conn:
            row = await conn.fetchrow("SELECT * FROM vendors WHERE id = $1", vendor_id)
            return self._row_to_dict(row) if row else None

    async def get_vendor_by_telegram(self, telegram_id: int) -> Optional[Dict[str, Any]]:
        """Fetch a vendor by their Telegram account id."""
        async with self._open_connection() as conn:
            row = await conn.fetchrow("SELECT * FROM vendors WHERE telegram_id = $1", telegram_id)
            return self._row_to_dict(row) if row else None

    async def create_vendor(self, telegram_id: int, name: str, menu: Optional[List[Dict[str, Any]]] = None) -> int:
        """Insert a new vendor and return its id."""
        menu_json = json.dumps(menu or [])
        async with self._open_connection() as conn:
            vendor_id = await conn.fetchval(
                "INSERT INTO vendors (telegram_id, name, menu_json) VALUES ($1, $2, $3) RETURNING id",
                telegram_id, name, menu_json
            )
            return int(vendor_id) if vendor_id else 0
        
    
    async def update_vendor(self, vendor_id: int, telegram_id: int, name: str, status: str = "active"):
        async with self._open_connection() as conn:
            await conn.execute(
                """
                UPDATE vendors
                SET telegram_id = $1,
                    name = $2,
                    status = $3,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = $4
                """,
                telegram_id, name, status, vendor_id
            )
        return vendor_id


    async def update_vendor_menu(self, vendor_id: int, menu: List[Dict[str, Any]]) -> None:
        """Update a vendor's menu JSON."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE vendors SET menu_json = $1, updated_at = CURRENT_TIMESTAMP WHERE id = $2",
                json.dumps(menu), vendor_id
            )

    async def set_vendor_status(self, vendor_id: int, status: str) -> None:
        """Activate/deactivate a vendor."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE vendors SET status = $1, updated_at = CURRENT_TIMESTAMP WHERE id = $2",
                status, vendor_id
            )

    # -------------------- Orders --------------------

    async def get_order_for_user(self, user_id: int) -> Optional[Dict[str, Any]]:
        """Returns the latest non-delivered order for a user."""
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM orders WHERE user_id = $1 AND status NOT IN ('delivered','cancelled') ORDER BY created_at DESC LIMIT 1",
                user_id
            )
            return self._row_to_dict(row) if row else None

    async def get_latest_active_order_for_dg(self, delivery_guy_id: int) -> Optional[Dict[str, Any]]:
        async with self._open_connection() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM orders WHERE delivery_guy_id = $1 AND status NOT IN ('delivered','cancelled') ORDER BY created_at DESC LIMIT 1",
                delivery_guy_id
            )
            return self._row_to_dict(row) if row else None

    # -------------------- Location Logs --------------------

    async def create_location_log(self, order_id: Optional[int], delivery_guy_id: Optional[int], lat: float, lon: float) -> int:
        async with self._open_connection() as conn:
            log_id = await conn.fetchval(
                "INSERT INTO location_logs (order_id, delivery_guy_id, lat, lon) VALUES ($1, $2, $3, $4) RETURNING id",
                order_id, delivery_guy_id, lat, lon
            )
            return int(log_id) if log_id else 0

    async def update_order_live(self, order_id: int, last_lat: float, last_lon: float):
        """Update the live location of the delivery guy for a specific order."""
        async with self._open_connection() as conn:
            await conn.execute(
                "UPDATE orders SET last_lat = $1, last_lon = $2, status = 'in_progress' WHERE id = $3",
                last_lat, last_lon, order_id
            )

    async def set_order_timestamp(self, order_id: int, field: str):
        assert field in ("accepted_at", "delivered_at")
        async with self._open_connection() as conn:
            await conn.execute(f"UPDATE orders SET {field} = CURRENT_TIMESTAMP WHERE id = $1", order_id)
   
   
    async def count_orders(
        self,
        filter_statuses: list[str] | None = None,
        delivery_guy_null: bool | None = None
    ) -> int:
        """
        Count orders with optional status filter and DG assignment filter.
        """
        async with self._open_connection() as conn:
            query = "SELECT COUNT(*) FROM orders WHERE TRUE"
            params = []
            param_index = 1

            if filter_statuses:
                query += f" AND status = ANY(${param_index})"
                params.append(filter_statuses)
                param_index += 1

            if delivery_guy_null is not None:
                query += f" AND (delivery_guy_id IS NULL) = ${param_index}"
                params.append(delivery_guy_null)
                param_index += 1

            return await conn.fetchval(query, *params) or 0
        
    async def count_active_delivery_guys(self) -> int:
        """
        Count all active, non-blocked delivery guys.
        """
        async with self._open_connection() as conn:
            return await conn.fetchval(
                "SELECT COUNT(*) FROM delivery_guys WHERE active = TRUE AND blocked = FALSE"
            ) or 0
        
    
    async def get_orders(
        self,
        filter_statuses: list[str] | None = None,
        delivery_guy_null: bool | None = None,
        limit: int = 10,
        offset: int = 0
    ) -> list[dict]:
        """
        Fetch orders with optional status filter and DG assignment filter.
        """
        async with self._open_connection() as conn:
            query = "SELECT * FROM orders WHERE TRUE"
            params = []
            param_index = 1

            if filter_statuses:
                query += f" AND status = ANY(${param_index})"
                params.append(filter_statuses)
                param_index += 1

            if delivery_guy_null is not None:
                query += f" AND (delivery_guy_id IS NULL) = ${param_index}"
                params.append(delivery_guy_null)
                param_index += 1

            query += f" ORDER BY created_at DESC LIMIT ${param_index} OFFSET ${param_index+1}"
            params.extend([limit, offset])

            rows = await conn.fetch(query, *params)
            return [dict(r) for r in rows]


    async def get_order(self, order_id: int) -> Optional[Dict[str, Any]]:
            async with self._open_connection() as conn:
                row = await conn.fetchrow("SELECT * FROM orders WHERE id = $1", order_id)
                return self._row_to_dict(row) if row else None
        
    
    async def update_order_status(self, order_id: int, status: str, dg_id: Optional[int] = None) -> None:
        """Updates the order status and handles time-based fields."""
        sql_parts = ["status = $1", "updated_at = CURRENT_TIMESTAMP"]
        params = [status]
        p = 2

        if dg_id:
            sql_parts.append(f"delivery_guy_id = ${p}")
            params.append(dg_id)
            p += 1

        if status in ("accepted", "preparing", "ready"):
            sql_parts.append("accepted_at = CURRENT_TIMESTAMP")
            if status == "ready":
                sql_parts.append("ready_at = CURRENT_TIMESTAMP")
        elif status == "delivered":
            sql_parts.append("delivered_at = CURRENT_TIMESTAMP")

        sql = f"UPDATE orders SET {', '.join(sql_parts)} WHERE id = ${p}"
        params.append(order_id)

        await self._pool.execute(sql, *params)


    async def update_order_delivery_guy(self, order_id: int, delivery_guy_id: int, breakdown_json: str | None) -> None:
        await self._pool.execute(
            """
            UPDATE orders
            SET delivery_guy_id = $1,
                breakdown_json = $2,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = $3
            """,
            delivery_guy_id, breakdown_json, order_id
        )


    async def set_order_timestamp(self, order_id: int, field_name: str) -> None:
        if field_name not in {"accepted_at", "ready_at", "delivered_at"}:
            raise ValueError(f"Unsupported timestamp field: {field_name}")
        await self._pool.execute(
            f"UPDATE orders SET {field_name} = CURRENT_TIMESTAMP WHERE id = $1",
            order_id
        )


    async def list_delivery_guys(
        self,
        limit: int,
        offset: int,
        active_only: bool = True
    ) -> list[dict]:
        """
        List delivery guys for manual assignment.
        """
        async with self._open_connection() as conn:
            query = """
            SELECT * FROM delivery_guys
            WHERE (active = TRUE OR $3 = FALSE)
            ORDER BY name
            LIMIT $1 OFFSET $2
            """
            rows = await conn.fetch(query, limit, offset, active_only)
            return [dict(r) for r in rows]


from datetime import date, datetime, timedelta
from typing import Dict, Any

class AnalyticsService:
    def __init__(self, db):
        self.db = db

    from datetime import datetime, timedelta
from typing import Dict, Any, List

class AnalyticsService:
    def __init__(self, db):
        self.db = db

    async def summarize_day(self) -> Dict[str, Any]:
        today = datetime.now().date()
        cutoff = datetime.now() - timedelta(days=1)

        # --- Users ---
        users = await self.db.get_leaderboard(limit=100)
        total_users = len(users)
        active_count = sum(
            1 for u in users
            if isinstance(u.get("last_active"), datetime) and u["last_active"] >= cutoff
        )
        new_users = await self.db.count_new_users(date=today)

        # --- Orders (delivered-only revenue now) ---
        orders_summary = await self.db.summarize_orders_day(date=today)
        orders_total = orders_summary["total"]
        orders_delivered = orders_summary["delivered"]
        orders_cancelled = orders_summary["cancelled"]
        food_rev = orders_summary["food_revenue"]        # delivered only
        delivery_fees = orders_summary["delivery_fees"]  # delivered only
        total_payout = food_rev + delivery_fees
        reliability_pct = orders_summary["reliability_pct"]

        # --- Cancelled names + top meal/vendor delivered/cancelled ---
        cancelled_orders: List[Dict[str, Any]] = await self.db.list_cancelled_orders_day(date=today)
        top_meal_name, top_meal_count = await self.db.top_meal_day(date=today)
        top_vendor_delivered, vendor_delivered_count = await self.db.top_vendor_delivered_day(date=today)
        top_vendor_cancelled, vendor_cancelled_count = await self.db.top_vendor_cancelled_day(date=today)

        # --- Vendors ---
        vendors_summary = await self.db.summarize_vendors_day(date=today)
        vendors_active = vendors_summary["active"]
        avg_vendor_rating = vendors_summary["avg_rating"] or 0

        # --- Delivery Guys ---
        dg_summary = await self.db.summarize_delivery_day(date=today)
        dg_active = dg_summary["active"]
        dg_deliveries = dg_summary["deliveries"]
        dg_acceptance_rate = dg_summary["acceptance_rate"]

        # --- Campus breakdown ---
        top_campus_name, top_campus_orders = await self.db.top_campus_day(date=today)

        # --- Top Deliverer (from leaderboard) ---
        top_deliverer_name = users[0].get("first_name", "N/A") if users else "None"
        top_deliverer_xp = users[0].get("xp", 0) if users else 0

        return {
            "date": today,
            "total_users": total_users,
            "active_count": active_count,
            "new_users": new_users,
            "orders_total": orders_total,
            "orders_delivered": orders_delivered,
            "orders_cancelled": orders_cancelled,
            "food_rev": food_rev,
            "delivery_fees": delivery_fees,
            "total_payout": total_payout,
            "vendors_active": vendors_active,
            "avg_vendor_rating": avg_vendor_rating,
            "dg_active": dg_active,
            "dg_deliveries": dg_deliveries,
            "dg_acceptance_rate": dg_acceptance_rate,
            "reliability_pct": reliability_pct,
            "top_campus_name": top_campus_name,
            "top_campus_orders": top_campus_orders,
            "top_deliverer_name": top_deliverer_name,
            "top_deliverer_xp": top_deliverer_xp,
            # New insights
            "cancelled_orders": cancelled_orders,
            "top_meal_name": top_meal_name,
            "top_meal_count": top_meal_count,
            "top_vendor_delivered": top_vendor_delivered,
            "vendor_delivered_count": vendor_delivered_count,
            "top_vendor_cancelled": top_vendor_cancelled,
            "vendor_cancelled_count": vendor_cancelled_count,
        }

    async def summary_text(self) -> str:
        data = await self.summarize_day()
        cancelled_names = ", ".join(
            [f"{o['meal_name']} ({o['vendor_name']})" for o in data["cancelled_orders"]]
        ) or "None"

        return f"""
📊⚡Daily Summary — {data['date']}
──────────────────────────────


👥 USERS: {data['total_users']}

    🔷 Active 24h: {data['active_count']}
    🆕 New: {data['new_users']}

📦 ORDERS: {data['orders_total']}

    💠 Delivered: {data['orders_delivered']}
    ❌ Cancelled: {data['orders_cancelled']}

💸 REVENUE: {data['total_payout']:.2f} ብር

    🍽️ Food: {data['food_rev']:.2f} ብር
    🚚 Delivery: {data['delivery_fees']:.2f} ብር

🍜 TOP MEAL: {data['top_meal_name']} ×{data['top_meal_count']}

🏪 VENDORS: {data['vendors_active']}

    ⭐ Avg Rating: {data['avg_vendor_rating']:.1f}
    🥇 Top Delivered: {data['top_vendor_delivered']} ({data['vendor_delivered_count']})
    🔻 Top Cancelled: {data['top_vendor_cancelled']} ({data['vendor_cancelled_count']})

🛵 DELIVERY SQUAD: {data['dg_active']}

    📡 Total Deliveries: {data['dg_deliveries']}
    📈 Acceptance Rate: {data['dg_acceptance_rate']:.0f}%

🧬 RELIABILITY INDEX: {data['reliability_pct']}%

🏛 TOP CAMPUS: {data['top_campus_name']} ({data['top_campus_orders']} orders)

🏆 TOP DELIVERER: {data['top_deliverer_name']} ({data['top_deliverer_xp']} XP)

⚡🧊 UniBites Delivery Bot • Powered by Neon Engine 🚀

              ..
    """
    
    async def delivery_report_text(self) -> str:
        from datetime import date
        today = date.today()

        async with self.db._pool.acquire() as conn:
            # Top drivers
            try:
                top_drivers = await self.db.get_top_drivers(today)
            except Exception:
                top_drivers = []

            # Low acceptance alerts
            driver_alerts = []
            rows = await conn.fetch(
                """
                SELECT DISTINCT ds.dg_id, dg.name
                FROM daily_stats ds
                LEFT JOIN delivery_guys dg ON dg.id = ds.dg_id
                WHERE ds.date = $1
                """,
                today
            )
            for r in rows:
                from utils.db_helpers import calc_acceptance_rate
                rate = await calc_acceptance_rate(self.db, r["dg_id"])
                if rate < 80.0:
                    driver_alerts.append(f"⚠️ {r['name']} • {rate:.1f}% acceptance")

            # Vendor cancels
            vendor_alerts = []
            cancel_rows = await conn.fetch(
                """
                SELECT v.name, COUNT(*) AS cancels
                FROM orders o
                JOIN vendors v ON o.vendor_id = v.id
                WHERE o.status = 'cancelled' AND o.created_at::DATE = $1
                GROUP BY v.name
                ORDER BY cancels DESC
                LIMIT 5
                """,
                today
            )
            vendor_alerts = [f"- {r['name']} • {int(r['cancels'])} cancels" for r in cancel_rows]

            # Engagement metric
            reactivated_count = int(await conn.fetchval(
                """
                SELECT COUNT(*) FROM delivery_guys
                WHERE last_online_at IS NOT NULL
                AND last_online_at > (NOW() - INTERVAL '2 hours')
                """
            ) or 0)

        admin_lines = [
            "━━━━━━━━━━━━━━━━━━━━━━",
            f"📢 **DELIVERY OPERATIONS DASHBOARD — {today}**",
            "━━━━━━━━━━━━━━━━━━━━━━",
            "🏆 **TOP PERFORMERS**:"
        ]
        if top_drivers:
            for idx, td in enumerate(top_drivers, 1):
                admin_lines.append(f"{idx}. **{td['name']}** — 🚚 {td['deliveries']} • 💵 {td['earnings']} birr")
        else:
            admin_lines.append("No top performers today.")

        admin_lines.append("")
        admin_lines.append("🚨 **ALERTS**")
        admin_lines.extend(driver_alerts or ["- No driver alerts"])
        admin_lines.extend(vendor_alerts or ["- No vendor alerts"])
        admin_lines.append("")
        admin_lines.append("📈 **ENGAGEMENT METRIC**")
        admin_lines.append(f"⚡ {reactivated_count} drivers bounced back online within 2 hours.")

        return "\n".join(admin_lines)
    
    async def summary_financial_text(self) -> str:
        today = datetime.now()
        header_date = today.strftime("%b %d, %Y · %A")  # "Dec 21, 2025 · Sunday"
        yesterday = today - timedelta(days=1)

        async with self.db._open_connection() as conn:
            # --- Today totals ---
            row_today = await conn.fetchrow("""
                SELECT 
                    SUM((breakdown_json::jsonb->>'vendor_share')::numeric) AS vendor_total,
                    SUM((breakdown_json::jsonb->>'platform_share')::numeric) AS admin_total
                FROM orders
                WHERE status='delivered' AND delivered_at::date = CURRENT_DATE
            """)
            vendor_total = row_today["vendor_total"] or 0
            admin_total = row_today["admin_total"] or 0

            # --- Yesterday totals ---
            row_yesterday = await conn.fetchrow("""
                SELECT 
                    SUM((breakdown_json::jsonb->>'vendor_share')::numeric) AS vendor_total,
                    SUM((breakdown_json::jsonb->>'platform_share')::numeric) AS admin_total
                FROM orders
                WHERE status='delivered' AND delivered_at::date = CURRENT_DATE - INTERVAL '1 day'
            """)
            vendor_total_y = row_yesterday["vendor_total"] or 0
            admin_total_y = row_yesterday["admin_total"] or 0

            # --- Commission tiers (today only) ---
            tiers = await conn.fetch("""
                SELECT (breakdown_json::jsonb->>'platform_share')::int AS commission, COUNT(*) AS count
                FROM orders
                WHERE status='delivered' AND delivered_at::date = CURRENT_DATE
                GROUP BY commission
                ORDER BY commission
            """)

            # --- Vendors (today only) ---
            vendors = await conn.fetch("""
                SELECT v.name,
                    COUNT(o.id) AS delivered_orders,
                    SUM(o.food_subtotal) AS gross_food,
                    SUM((o.breakdown_json::jsonb->>'platform_share')::numeric) AS commission,
                    SUM((o.breakdown_json::jsonb->>'vendor_share')::numeric) AS vendor_payout
                FROM orders o
                JOIN vendors v ON o.vendor_id = v.id
                WHERE o.status='delivered' AND o.delivered_at::date = CURRENT_DATE
                GROUP BY v.name
                ORDER BY vendor_payout DESC;
            """)

        # --- Format tier lines ---
        tier_lines = "\n".join(
        [f"      ▸ {(t['commission'] or 0)} birr tier: *{t['count']} orders*" for t in tiers]
    ) or "       ▸ None"

        # --- Format vendor lines ---
        vendor_lines = "\n".join(
        [
            f"   • {v['name']}: Gross {(v['gross_food'] or 0):.2f} ብር | "
            f"*Net {(v['vendor_payout'] or 0):.2f} ብር ({v['delivered_orders'] or 0} orders)*"
            for v in vendors
        ]
    ) or "   • None"

        # --- Gross totals ---
        gross_total = vendor_total + admin_total
        gross_total_y = vendor_total_y + admin_total_y

        # --- Percentage change helper ---
        def pct_change(today_val, yesterday_val):
            if yesterday_val == 0:
                return "N/A"
            diff = today_val - yesterday_val
            pct = (diff / yesterday_val) * 100
            arrow = "⬆️" if diff > 0 else ("⬇️" if diff < 0 else "➡️")
            return f"{arrow} {pct:.1f}% vs yesterday"

        vendor_trend = pct_change(vendor_total, vendor_total_y)
        admin_trend = pct_change(admin_total, admin_total_y)
        gross_trend = pct_change(gross_total, gross_total_y)

        return f"""
━━━━━━━━━━━━━━━━━━━━━━
💰⚡ *Financial Summary*
            *{header_date}*
━━━━━━━━━━━━━━━━━━━━━━

🏪 *VENDORS*

    • Total Vendor Revenue: *{vendor_total:.2f} ብር* 
    ({vendor_trend})
    
    {vendor_lines}

🛡 *ADMIN PROFIT*

    • Total Commission: *{admin_total:.2f} ብር* 
    ({admin_trend})
    
    {tier_lines}
    
    • Net Admin Profit: *{admin_total:.2f} ብር*

📦 *GROSS REVENUE*

    • Total Orders Revenue: *{gross_total:.2f} ብር* 
    ({gross_trend})
    
    • Vendor Payouts: *{vendor_total:.2f} ብር*
    • Admin Profit: *{admin_total:.2f} ብር*

⚡ *UniBites Delivery — transparent payouts, clear profits 🚀*
            """
            
        
    
    # ---------- Weekly summary core ----------
    async def summarize_week(self,
                            start_date: Optional[date] = None,
                            end_date: Optional[date] = None) -> Dict[str, Any]:
        """
        Summarize the 7-day window [start_date .. end_date] inclusive.
        Defaults: previous 7 days ending yesterday (i.e. yesterday and the 6 days before).
        Returns a dict with lots of metrics and a 'vendor_payouts' list with per-vendor details.
        """
        # default window: previous 7 days (yesterday inclusive)
        today = datetime.now().date()
        if end_date is None:
            end_date = today
        if start_date is None:
            start_date = date.min # 7-day window

        # # previous week window for comparisons
        # prev_end = start_date - timedelta(days=0)
        # prev_start = prev_end - timedelta(days=7)
        prev_end = start_date 
        prev_start = date.min

        # normalize to date strings if needed for SQL
        # We'll use DB-level aggregation (faster) where possible
        async with self.db._open_connection() as conn:
            # --- Users ---
            total_users = int(await conn.fetchval("SELECT COUNT(*) FROM users") or 0)
            new_users = int(await conn.fetchval(
                "SELECT COUNT(*) FROM users WHERE created_at::date BETWEEN $1 AND $2",
                start_date, end_date
            ) or 0)

            # Weekly Active Users (WAU): users with activity or last_active in the window
            wau = int(await conn.fetchval(
    """
    SELECT COUNT(DISTINCT user_id)
    FROM orders
    WHERE created_at::date BETWEEN $1 AND $2
    """,
    start_date, end_date
) or 0)
            
            prev_wau = int(await conn.fetchval(
    """
    SELECT COUNT(DISTINCT user_id)
    FROM orders
    WHERE created_at::date BETWEEN $1 AND $2
    """,
    prev_start, prev_end
    ) or 0)


            prev_new_users = int(await conn.fetchval(
                """
                SELECT COUNT(*)
                FROM users
                WHERE created_at::date BETWEEN $1 AND $2
                """,
                prev_start, prev_end
            ) or 0)



            # --- Orders & revenue (delivered only for revenue) ---
            orders_row = await conn.fetchrow(
                """
                SELECT
                    COUNT(*) FILTER (WHERE created_at::date BETWEEN $1 AND $2) AS total_orders,
                    COUNT(*) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS delivered,
                    COUNT(*) FILTER (WHERE status='cancelled' AND created_at::date BETWEEN $1 AND $2) AS cancelled,
                    SUM(food_subtotal) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS food_revenue,
                    SUM((breakdown_json::jsonb->>'platform_share')::numeric) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS delivery_fees
                FROM orders
                """,
                start_date, end_date
            )

            orders_total = int(orders_row["total_orders"] or 0)
            orders_delivered = int(orders_row["delivered"] or 0)
            orders_cancelled = int(orders_row["cancelled"] or 0)
            food_rev = Decimal(orders_row["food_revenue"] or 0)
            delivery_fees = Decimal(orders_row["delivery_fees"] or 0)
            total_payout = food_rev + delivery_fees
            reliability_pct = round((orders_delivered / orders_total * 100), 1) if orders_total > 0 else 0.0

            # previous week totals (for WoW)
            prev_row = await conn.fetchrow(
                """
                SELECT
                    COUNT(*) FILTER (WHERE created_at::date BETWEEN $1 AND $2) AS total_orders,
                    COUNT(*) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS delivered,
                    SUM(food_subtotal) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS food_revenue,
                    SUM((breakdown_json::jsonb->>'platform_share')::numeric) FILTER (WHERE status='delivered' AND delivered_at::date BETWEEN $1 AND $2) AS delivery_fees
                FROM orders
                """,
                prev_start, prev_end
            )
            prev_total_orders = int(prev_row["total_orders"] or 0)
            prev_delivered = int(prev_row["delivered"] or 0)
            prev_food_rev = Decimal(prev_row["food_revenue"] or 0)
            prev_delivery_fees = Decimal(prev_row["delivery_fees"] or 0)
            prev_total_payout = prev_food_rev + prev_delivery_fees

            # --- Cancelled orders list (basic info) ---
            cancelled_orders_rows = await conn.fetch(
                    """
                    SELECT v.name AS vendor_name, COUNT(*) AS cancelled_count
                    FROM orders o
                    JOIN vendors v ON v.id = o.vendor_id
                    WHERE o.status = 'cancelled'
                    AND o.created_at::date BETWEEN $1 AND $2
                    GROUP BY v.name
                    ORDER BY cancelled_count DESC
                    LIMIT 5
                    """,
                    start_date, end_date
                )
            cancelled_orders = [
                {"vendor_name": r["vendor_name"], "cancelled_count": int(r["cancelled_count"])  }
                for r in cancelled_orders_rows
            ]

            # --- Top meal & campus ---    
            rows = await conn.fetch(
                """
                SELECT items_json
                FROM orders
                WHERE created_at::date BETWEEN $1 AND $2
                """,
                start_date, end_date
            )

            meal_counter = Counter()
            for row in rows:
                items = json.loads(row["items_json"] or "[]")
                for item in items:
                    meal_counter[item["name"]] += item.get("qty", 1)

            if meal_counter:
                top_meal_name, top_meal_count = meal_counter.most_common(1)[0]
            else:
                top_meal_name, top_meal_count = "N/A", 0

            top_campus_row = await conn.fetchrow(
                """
                SELECT u.campus AS campus_name, COUNT(*) AS cnt
                FROM orders o
                JOIN users u ON u.id = o.user_id
                WHERE o.created_at::date BETWEEN $1 AND $2
                AND u.campus IS NOT NULL
                GROUP BY u.campus
                ORDER BY cnt DESC
                LIMIT 1
                """,
                start_date, end_date
            )

            top_campus_name = top_campus_row["campus_name"] if top_campus_row else "None"
            top_campus_orders = int(top_campus_row["cnt"]) if top_campus_row else 0
            # --- Vendors summary & per-vendor payouts (delivered only) ---
            vendors_rows = await conn.fetch(
                """
                SELECT v.id, v.name,
                    COUNT(o.id) FILTER (WHERE o.status='delivered' AND o.delivered_at::date BETWEEN $1 AND $2) AS delivered_orders,
                    SUM(o.food_subtotal) FILTER (WHERE o.status='delivered' AND o.delivered_at::date BETWEEN $1 AND $2) AS gross_food,
                    SUM((o.breakdown_json::jsonb->>'platform_share')::numeric) FILTER (WHERE o.status='delivered' AND o.delivered_at::date BETWEEN $1 AND $2) AS commission,
                    SUM((o.breakdown_json::jsonb->>'vendor_share')::numeric) FILTER (WHERE o.status='delivered' AND o.delivered_at::date BETWEEN $1 AND $2) AS vendor_payout,
                    COUNT(o.id) FILTER (WHERE o.status='cancelled' AND o.created_at::date BETWEEN $1 AND $2) AS cancelled
                FROM vendors v
                LEFT JOIN orders o ON o.vendor_id = v.id
                GROUP BY v.id, v.name
                ORDER BY vendor_payout DESC
                LIMIT 50
                """,
                start_date, end_date
            )

            vendor_payouts = []
            for r in vendors_rows:
                vendor_payouts.append({
                    "id": r["id"],
                    "name": r["name"],
                    "delivered_orders": int(r["delivered_orders"] or 0),
                    "gross_food": Decimal(r["gross_food"] or 0),
                    "commission": Decimal(r["commission"] or 0),
                    "vendor_payout": Decimal(r["vendor_payout"] or 0),
                    "cancelled": int(r["cancelled"] or 0),
                    "reliability": round(
                        (int(r["delivered_orders"] or 0) / ((int(r["delivered_orders"] or 0) + int(r["cancelled"] or 0)) or 1)) * 100,
                        1
                    )
                })

            # vendor-level aggregated totals
            vendors_active = int(await conn.fetchval(
    """
    SELECT COUNT(DISTINCT o.vendor_id)
    FROM orders o
    WHERE o.created_at::date BETWEEN $1 AND $2
    """,
    start_date, end_date
) or 0)
            avg_vendor_rating = float(await conn.fetchval("SELECT AVG(rating_avg) FROM vendors") or 0.0)

            # --- Delivery guys summary (weekly aggregates) ---
            dg_row = await conn.fetchrow(
                """
                SELECT
                    COUNT(DISTINCT dg_id) FILTER (WHERE ds.date BETWEEN $1 AND $2) AS active_dg,
                    SUM(ds.deliveries) FILTER (WHERE ds.date BETWEEN $1 AND $2) AS total_deliveries
                FROM daily_stats ds
                """,
                start_date, end_date
            )
            dg_active = int(dg_row["active_dg"] or 0)
            dg_deliveries = int(dg_row["total_deliveries"] or 0)

            # approximate acceptance rate: compute from daily_stats averages if available
            acc_row = await conn.fetchrow(
                """
                SELECT AVG(ds.acceptance_rate) AS avg_acc
                FROM daily_stats ds
                WHERE ds.date BETWEEN $1 AND $2
                """,
                start_date, end_date
            )
            dg_acceptance_rate = float(acc_row["avg_acc"] or 0.0)

            # --- Top deliverer from leaderboard (reuse existing helper if present) ---
            leaderboard = await self.db.get_leaderboard(limit=200)
            top_deliverer_name = leaderboard[0].get("first_name") if leaderboard else "None"
            top_deliverer_xp = leaderboard[0].get("xp") if leaderboard else 0

            # --- compute WoW deltas for key KPIs ---
            def pct_change_val(curr: Decimal, prev: Decimal) -> Tuple[str, float]:
                if prev == 0:
                    return ("N/A", 0.0)
                diff = (curr - prev)
                pct = float((diff / prev) * 100)
                arrow = "up" if diff > 0 else ("down" if diff < 0 else "flat")
                return (arrow, round(pct, 1))

            vendor_total = sum(v["vendor_payout"] for v in vendor_payouts)
            admin_total = delivery_fees  # your system uses delivery_fees as platform/commission here
            gross_total = vendor_total + admin_total

            prev_vendor_total = prev_total_payout  # best approximation using earlier query
            prev_admin_total = prev_delivery_fees
            prev_gross_total = prev_total_payout

            vendor_trend = pct_change_val(Decimal(vendor_total), Decimal(prev_vendor_total))
            admin_trend = pct_change_val(Decimal(admin_total), Decimal(prev_admin_total))
            gross_trend = pct_change_val(Decimal(gross_total), Decimal(prev_gross_total))

        # --- Generate automated action recommendations ---
        recommendations = self._generate_weekly_recommendations({
            "orders_total": orders_total,
            "orders_delivered": orders_delivered,
            "orders_cancelled": orders_cancelled,
            "cancelled_orders": cancelled_orders,
            "vendor_payouts": vendor_payouts,
            "dg_acceptance_rate": dg_acceptance_rate,
            "vendor_total": float(vendor_total),
            "vendor_trend_pct": vendor_trend[1] if vendor_trend[0] != "N/A" else None,
        }, prev_metrics={
            "orders_total": prev_total_orders,
            "orders_delivered": prev_delivered,
            "vendor_total": float(prev_vendor_total)
        })

        # Compose summary dict
        return {
            "start_date": start_date,
            "end_date": end_date,
            "total_users": total_users,
            "wau": wau,
            "new_users": new_users,
            "prev_new_users": prev_new_users,
            "prev_wau": prev_wau,
            "orders_total": orders_total,
            "orders_delivered": orders_delivered,
            "orders_cancelled": orders_cancelled,
            "food_rev": float(food_rev),
            "delivery_fees": float(delivery_fees),
            "total_payout": float(total_payout),
            "vendors_active": vendors_active,
            "avg_vendor_rating": avg_vendor_rating,
            "vendor_payouts": vendor_payouts,
            "vendor_total": float(vendor_total),
            "admin_total": float(admin_total),
            "gross_total": float(gross_total),
            "vendor_trend": vendor_trend,
            "admin_trend": admin_trend,
            "gross_trend": gross_trend,
            "dg_active": dg_active,
            "dg_deliveries": dg_deliveries,
            "dg_acceptance_rate": dg_acceptance_rate,
            "top_campus_name": top_campus_name,
            "top_campus_orders": top_campus_orders,
            "top_meal_name": top_meal_name,
            "top_meal_count": top_meal_count,
            "top_deliverer_name": top_deliverer_name,
            "top_deliverer_xp": top_deliverer_xp,
            "reliability_pct": reliability_pct,
            "cancelled_orders": cancelled_orders,
            "recommendations": recommendations,
        }

    # ---------- Recommendations (simple rule-based) ----------
    def _generate_weekly_recommendations(self, metrics: Dict[str, Any], prev_metrics: Dict[str, Any]) -> List[str]:
        """
        Return a list of short, actionable recommendations based on weekly metrics.
        This is intentionally rule-based and explainable.
        """
        recs: List[str] = []

        # cancellations
        canc_rate = (metrics["orders_cancelled"] / metrics["orders_total"] * 100) if metrics["orders_total"] else 0.0
        if canc_rate >= 10:
            recs.append(f"High cancellation rate this week ({canc_rate:.1f}%). Investigate top cancelled meals/vendors and contact them.")

        # vendor reliability
        low_reliability_vendors = [v for v in metrics["vendor_payouts"] if v["reliability"] < 85]
        if low_reliability_vendors:
            names = ", ".join(v["name"] for v in low_reliability_vendors[:3])
            recs.append(f"Vendors with low reliability: {names}. Consider warnings, training, or temporary delisting.")

        # vendor payouts drop
        prev_vendor_total = prev_metrics.get("vendor_total", 0) or 0
        if prev_vendor_total and metrics["vendor_total"] < prev_vendor_total * 0.9:
            recs.append("Vendor payouts down >10% vs previous week — consider promotions or vendor-side issues (stock/outages).")

        # low DG acceptance
        if metrics["dg_acceptance_rate"] < 75:
            recs.append(f"DG acceptance low ({metrics['dg_acceptance_rate']:.1f}%). Consider incentives during peaks or recruit more DGs.")

        # repeat offenders (lots of cancels)
        offenders = sorted(metrics["vendor_payouts"], key=lambda v: v["cancelled"], reverse=True)
        if offenders and offenders[0]["cancelled"] >= 10:
            recs.append(f"Top canceller: {offenders[0]['name']} with {offenders[0]['cancelled']} cancels. Investigate immediately.")

        # growth opportunity
        if metrics["orders_total"] >= 200 and metrics["dg_acceptance_rate"] > 80:
            recs.append("Good demand this week. Consider a small marketing push to convert new users to repeat customers.")

        if not recs:
            recs.append("No urgent actions detected — system is stable. Keep monitoring next week.")

        return recs

    # ---------- Weekly text formatter ----------
    async def summary_week_text(self,
                                start_date: Optional[date] = None,
                                end_date: Optional[date] = None) -> str:
        """
        Build a Telegram-friendly string summary for the week using summarize_week().
        """
        data = await self.summarize_week(start_date=start_date, end_date=end_date)
        sd = data["start_date"].strftime("%b %d")
        ed = data["end_date"].strftime("%b %d, %Y")

        # build vendor lines (top 8)
        vendor_lines = []
        for v in data["vendor_payouts"][:8]:
            vendor_lines.append(
                f"• {v['name']}: Gross {v['gross_food']:.2f} ብር | Net {v['vendor_payout']:.2f} ብር | Cancels {v['cancelled']} | Reliability {v['reliability']}%\n"
            )
        vendor_lines_text = "\n".join(vendor_lines) or "• None"

        canceled_preview = []

        for o in data["cancelled_orders"][:6]:
            items = json.loads(o.get("items_json") or "[]")
            if items:
                # Take first item's name for preview
                meal_name = items[0]["name"]
            else:
                meal_name = "N/A"
            vendor_name = o.get("vendor_name", "Unknown")
            canceled_preview.append(f"{meal_name}({vendor_name})")

        canceled_preview = ", ".join(canceled_preview) or "None"

        recs_text = "\n\n".join([f"• {r}" for r in data["recommendations"][:6]])

        return f"""
━━━━━━━━━━━━━━━━━━━━━━
📊⚡ *Weekly Summary — {sd} → {ed}*
━━━━━━━━━━━━━━━━━━━━━━

👥 *USERS*
    • Total users: *{data['total_users']}*
    • Weekly Active (WAU): *{data['wau']}*
    • New users this week: *{data['new_users']}*

📦 *ORDERS*
    • Total orders: *{data['orders_total']}*
    • Delivered: *{data['orders_delivered']}*
    • Cancelled: *{data['orders_cancelled']}*
    • Reliability: *{data['reliability_pct']:.1f}%*

💸 *FINANCIALS*
    • Vendor payouts (total): *{data['vendor_total']:.2f} ብር*
    • Admin (commission): *{data['admin_total']:.2f} ብር*
    • Gross revenue: *{data['gross_total']:.2f} ብር*

🏪 *TOP VENDORS (sample)*
{vendor_lines_text}  # pre-formatted vendor lines, max 4-5 vendors

🍜 *TOP MEAL*: *{data['top_meal_name']}* ×{data['top_meal_count']}
📍 *TOP CAMPUS*: *{data['top_campus_name']}* ({data['top_campus_orders']} orders)

🛵 *DELIVERY SQUAD*
    • Active DGs: *{data['dg_active']}*
    • Total deliveries: *{data['dg_deliveries']}*
    • Acceptance rate: *{data['dg_acceptance_rate']:.1f}%*

❌ *CANCELLATIONS (sample)*
{canceled_preview}  # top cancelled meals/vendors

🧠 *RECOMMENDATIONS*
{recs_text}  # list of recommendations for admin actions

⚡ *UniBites Delivery — weekly ops snapshot 🚀*
━━━━━━━━━━━━━━━━━━━━━━
"""


    # ---------- PDF export ----------
   
    def export_weekly_pdf(self, summary: Dict[str, Any], path: str) -> None:
        """
        Export a polished Excel file of the weekly summary.
        summary: result of summarize_week()
        path: filesystem path to write Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Summary"

        # ---------- Styles ----------
        header_font = Font(bold=True, size=14, color="FFFFFF")
        section_font = Font(bold=True, size=11)
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        blue_fill = PatternFill("solid", fgColor="4F81BD")
        light_gray = PatternFill("solid", fgColor="F2F2F2")

        # Named style for table headers
        header_style = NamedStyle(name="table_header")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill("solid", fgColor="2F5597")
        header_style.alignment = center
        header_style.border = border
        if "table_header" not in wb.named_styles:
            wb.add_named_style(header_style)

        # ---------- Title ----------
        start = summary.get("start_date")
        end = summary.get("end_date")
        start_str = start.strftime("%b %d") if hasattr(start, "strftime") else str(start)
        end_str = end.strftime("%b %d, %Y") if hasattr(end, "strftime") else str(end)
        ws.merge_cells("A1:F1")
        ws["A1"] = f"📊 Weekly Summary • {start_str} → {end_str}"
        ws["A1"].font = header_font
        ws["A1"].alignment = center
        ws["A1"].fill = blue_fill
        ws.row_dimensions[1].height = 28

        row = 3

        # ---------- USERS ----------
        ws[f"A{row}"] = "👥 USERS"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Total users"
        ws[f"B{row}"] = summary.get("total_users", 0)
        ws[f"A{row}"].font = bold_font
        row += 1
        ws[f"A{row}"] = "WAU"
        ws[f"B{row}"] = summary.get("wau", 0)
        ws[f"A{row}"].font = bold_font
        row += 1
        ws[f"A{row}"] = "New users this week"
        ws[f"B{row}"] = summary.get("new_users", 0)
        row += 2

        # ---------- ORDERS ----------
        ws[f"A{row}"] = "📦 ORDERS"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Total orders"
        ws[f"B{row}"] = summary.get("orders_total", 0)
        ws[f"C{row}"] = "Delivered"
        ws[f"D{row}"] = summary.get("orders_delivered", 0)
        ws[f"E{row}"] = "Cancelled"
        ws[f"F{row}"] = summary.get("orders_cancelled", 0)
        for col in ("A","C","E"):
            ws[f"{col}{row}"].font = bold_font
        row += 1
        ws[f"A{row}"] = "Reliability (%)"
        ws[f"B{row}"] = summary.get("reliability_pct", 0.0)
        ws[f"B{row}"].number_format = '0.00%'
        row += 2

        # ---------- FINANCIALS ----------
        ws[f"A{row}"] = "💸 FINANCIALS"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Vendor payouts"
        ws[f"B{row}"] = summary.get("vendor_total", 0.0)
        ws[f"B{row}"].number_format = '#,##0.00 "birr"'
        ws[f"C{row}"] = "Admin commission"
        ws[f"D{row}"] = summary.get("admin_total", 0.0)
        ws[f"D{row}"].number_format = '#,##0.00 "birr"'
        ws[f"E{row}"] = "Gross revenue"
        ws[f"F{row}"] = summary.get("gross_total", 0.0)
        ws[f"F{row}"].number_format = '#,##0.00 "birr"'
        for col in ("A","C","E"):
            ws[f"{col}{row}"].font = bold_font
        row += 2

        # ---------- TOP VENDORS TABLE ----------
        ws[f"A{row}"] = "🏪 TOP VENDORS"
        ws[f"A{row}"].font = section_font
        row += 1
        vendor_header = ["Vendor", "Net Payout", "Gross Food", "Cancels"]
        ws.append(vendor_header)
        header_row = ws.max_row
        for col_idx in range(1, len(vendor_header) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.style = "table_header"

        for v in summary.get("vendor_payouts", [])[:20]:
            ws.append([
                v.get("name", "Unknown"),
                v.get("vendor_payout", 0.0),
                v.get("gross_food", 0.0),
                v.get("cancelled", 0)
            ])
        # format numeric columns
        start_data_row = header_row + 1
        end_data_row = ws.max_row
        for r in range(start_data_row, end_data_row + 1):
            ws[f"B{r}"].number_format = '#,##0.00 "birr"'
            ws[f"C{r}"].number_format = '#,##0.00 "birr"'
            ws[f"D{r}"].number_format = '0'

        # Add a bar chart for top vendors (if data present)
        if end_data_row >= start_data_row:
            chart = BarChart()
            chart.title = "Top Vendor Net Payouts"
            chart.y_axis.title = "Net Payout (birr)"
            chart.x_axis.title = "Vendor"
            data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=end_data_row)
            cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=end_data_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 8
            chart.width = 14
            chart_anchor_row = end_data_row + 2
            ws.add_chart(chart, f"A{chart_anchor_row}")

        row = ws.max_row + 3

        # ---------- TOP MEAL & CAMPUS ----------
        ws[f"A{row}"] = "🍜 TOP MEAL"
        ws[f"A{row}"].font = section_font
        ws[f"B{row}"] = f"{summary.get('top_meal_name','N/A')} ×{summary.get('top_meal_count',0)}"
        row += 1
        ws[f"A{row}"] = "📍 TOP CAMPUS"
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"] = f"{summary.get('top_campus_name','N/A')} ({summary.get('top_campus_orders',0)} orders)"
        row += 2

        # ---------- DELIVERY SQUAD ----------
        ws[f"A{row}"] = "🛵 DELIVERY SQUAD"
        ws[f"A{row}"].font = section_font
        row += 1
        ws.append(["Active DGs", "Total Deliveries", "Acceptance Rate"])
        ds_header_row = ws.max_row
        for col_idx in range(1, 4):
            ws.cell(row=ds_header_row, column=col_idx).style = "table_header"
        ws.append([
            summary.get("dg_active", 0),
            summary.get("dg_deliveries", 0),
            summary.get("dg_acceptance_rate", 0.0)
        ])
        # format acceptance rate
        ws[f"C{ds_header_row + 1}"].number_format = '0.00%'

        row = ws.max_row + 2

        # ---------- CANCELLATIONS ----------
        ws[f"A{row}"] = "❌ CANCELLATIONS (sample)"
        ws[f"A{row}"].font = section_font
        row += 1
        cancelled_preview = summary.get("cancelled_preview", "N/A")
        ws[f"A{row}"] = cancelled_preview
        row += 2

        # ---------- RECOMMENDATIONS ----------
        ws[f"A{row}"] = "🧠 RECOMMENDATIONS"
        ws[f"A{row}"].font = section_font
        row += 1
        for r in summary.get("recommendations", []):
            ws[f"A{row}"] = r
            row += 1

        # ---------- Footer / Prepared on ----------
        footer_row = ws.max_row + 2
        ws[f"E{footer_row}"] = f"Prepared on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws[f"E{footer_row}"].alignment = Alignment(horizontal="right")
        ws[f"E{footer_row}"].font = Font(italic=True, size=9)

        # ---------- Layout polish ----------
        # Freeze top rows for easy navigation
        ws.freeze_panes = "A4"

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                except Exception:
                    val = ""
                if len(val) > max_length:
                    max_length = len(val)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width if adjusted_width < 60 else 60

        # Conditional formatting examples
        # Highlight reliability below 80% (if cell exists)
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            if isinstance(cell.value, (int, float)):
                # apply a rule to the whole column B for rows where reliability might appear
                pass
        # Apply conditional formatting to the reliability cell if present
        # Find reliability cell (search for label)
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Reliability (%)":
                rel_cell = f"B{r}"
                ws.conditional_formatting.add(rel_cell,
                    CellIsRule(operator='lessThan', formula=['0.8'], stopIfTrue=True,
                            fill=PatternFill("solid", fgColor="FFC7CE")))
                break

        # Save Excel
        wb.save(path)

        # returns nothing; file is written to path

    # ---------- Helper: pretty percent ----------
    @staticmethod
    def _fmt_pct(v: float) -> str:
        try:
            return f"{v:.1f}%"
        except Exception:
            return "N/A"

# -------------------- Seed Functions --------------------
async def seed_vendors(db: Database) -> None:
    vendors = [
        {
    "telegram_id": 589745233,
    "name": "ቲጂ አቡዳቢ #5kilo",
    "menu": [
     {"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
                {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
                {"id": 78, "name": "Energy Drink", "price": 70, "category": "Drinks"},
                {"id": 79, "name": "Coca SD", "price": 75, "category": "Drinks"},
                {"id": 80, "name": "Sprite SD", "price": 75, "category": "Drinks"},
                {"id": 81, "name": "Fanta p.apple SD", "price": 75, "category": "Drinks"},
        {"id": 1, "name": "ሃፍ ሃፍ", "price": 160, "category": "Fasting"},
        {"id": 2, "name": "ፓስታ በስጎ", "price": 130, "category": "Fasting"},
        {"id": 3, "name": "ፓስታ በአትክልት", "price": 130, "category": "Fasting"},
        {"id": 4, "name": "ፍርፍር በቀይ", "price": 130, "category": "Fasting"},
        {"id": 5, "name": "ፍርፍር በአልጫ", "price": 130, "category": "Fasting"},
        {"id": 6, "name": "ግማሽ ኮርኒስ", "price": 160, "category": "Fasting"},
        {"id": 7, "name": "ሙሉ ኮርኒስ", "price": 260, "category": "Fasting"},
        {"id": 8, "name": "ሙሉ አገልግል", "price": 290, "category": "Fasting"},
        {"id": 9, "name": "ግማሽ አገልገል", "price": 180, "category": "Fasting"},
        {"id": 10, "name": "በየዓይነት", "price": 140, "category": "Fasting"},
        {"id": 11, "name": "ቴስቲ ወጥ", "price": 130, "category": "Fasting"},
        {"id": 12, "name": "ቴስቲ ጥብስ", "price": 140, "category": "Fasting"},
        {"id": 13, "name": "ስፔሻል ፍርፍር", "price": 160, "category": "Fasting"},
        {"id": 14, "name": "ስጋ ፍርፍር", "price": 190, "category": "Non Fasting"},
        {"id": 15, "name": "ጥብስ ፍርፍር", "price": 230, "category": "Non Fasting"},
        {"id": 16, "name": "እንቁላል ፍርፍር", "price": 160, "category": "Non Fasting"},
        {"id": 17, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
        {"id": 18, "name": "ሙሉ ኮርኒስ (የፍስክ)", "price": 360, "category": "Non Fasting"},
        {"id": 19, "name": "ግማሽ ኮርኒስ (የፍስክ)", "price": 210, "category": "Non Fasting"},
        {"id": 20, "name": "ፓስታ በእንቁላል", "price": 160, "category": "Non Fasting"},
        {"id": 21, "name": "ፓስታ በስጋ", "price": 190, "category": "Non Fasting"},
        {"id": 22, "name": "ስፔሻል ፍርፍር", "price": 260, "category": "Non Fasting"},
        {"id": 23, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
        {"id": 24, "name": "ሙሉ አገልግል (የፍስክ)", "price": 410, "category": "Non Fasting"},
        {"id": 25, "name": "ግማሽ አገልግል(የፍስክ)", "price": 260, "category": "Non Fasting"},
        {"id": 26, "name": "አይብ", "price": 210, "category": "Non Fasting"},
        {"id": 27, "name": "ስፔሻል ኮርኒስ (የፍስክ)", "price": 410, "category": "Specials"}
    ]
    },

        {
        "telegram_id": 6567214347,
            "name": "ቲጂ አቡዳቢ #6kilo",
            "menu": [
      {"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
                {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
                {"id": 78, "name": "Energy Drink", "price": 70, "category": "Drinks"},
                {"id": 79, "name": "Coca SD", "price": 75, "category": "Drinks"},
                {"id": 80, "name": "Sprite SD", "price": 75, "category": "Drinks"},
                {"id": 81, "name": "Fanta p.apple SD", "price": 75, "category": "Drinks"},
        {"id": 1, "name": "ሃፍ ሃፍ", "price": 160, "category": "Fasting"},
        {"id": 2, "name": "ፓስታ በስጎ", "price": 130, "category": "Fasting"},
        {"id": 3, "name": "ፓስታ በአትክልት", "price": 130, "category": "Fasting"},
        {"id": 4, "name": "ፍርፍር በቀይ", "price": 130, "category": "Fasting"},
        {"id": 5, "name": "ፍርፍር በአልጫ", "price": 130, "category": "Fasting"},
        {"id": 6, "name": "ግማሽ ኮርኒስ", "price": 160, "category": "Fasting"},
        {"id": 7, "name": "ሙሉ ኮርኒስ", "price": 260, "category": "Fasting"},
        {"id": 8, "name": "ሙሉ አገልግል", "price": 290, "category": "Fasting"},
        {"id": 9, "name": "ግማሽ አገልገል", "price": 180, "category": "Fasting"},
        {"id": 10, "name": "በየዓይነት", "price": 140, "category": "Fasting"},
        {"id": 11, "name": "ቴስቲ ወጥ", "price": 130, "category": "Fasting"},
        {"id": 12, "name": "ቴስቲ ጥብስ", "price": 140, "category": "Fasting"},
        {"id": 13, "name": "ስፔሻል ፍርፍር", "price": 160, "category": "Fasting"},
        {"id": 14, "name": "ስጋ ፍርፍር", "price": 190, "category": "Non Fasting"},
        {"id": 15, "name": "ጥብስ ፍርፍር", "price": 230, "category": "Non Fasting"},
        {"id": 16, "name": "እንቁላል ፍርፍር", "price": 160, "category": "Non Fasting"},
        {"id": 17, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
        {"id": 18, "name": "ሙሉ ኮርኒስ (የፍስክ)", "price": 360, "category": "Non Fasting"},
        {"id": 19, "name": "ግማሽ ኮርኒስ (የፍስክ)", "price": 210, "category": "Non Fasting"},
        {"id": 20, "name": "ፓስታ በእንቁላል", "price": 160, "category": "Non Fasting"},
        {"id": 21, "name": "ፓስታ በስጋ", "price": 190, "category": "Non Fasting"},
        {"id": 22, "name": "ስፔሻል ፍርፍር", "price": 260, "category": "Non Fasting"},
        {"id": 23, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
        {"id": 24, "name": "ሙሉ አገልግል (የፍስክ)", "price": 410, "category": "Non Fasting"},
        {"id": 25, "name": "ግማሽ አገልግል(የፍስክ)", "price": 260, "category": "Non Fasting"},
        {"id": 26, "name": "አይብ", "price": 210, "category": "Non Fasting"},
        {"id": 27, "name": "ስፔሻል ኮርኒስ (የፍስክ)", "price": 410, "category": "Specials"}
    ],
        },
       {
  "telegram_id": 8487056502,
  "name": "ጤና ምግብ ቤት",
  "menu": [
{"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
                {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
                {"id": 78, "name": "Energy Drink", "price": 70, "category": "Drinks"},
                {"id": 79, "name": "Coca SD", "price": 75, "category": "Drinks"},
                {"id": 80, "name": "Sprite SD", "price": 75, "category": "Drinks"},
                {"id": 81, "name": "Fanta p.apple SD", "price": 75, "category": "Drinks"},
    {"id": 1, "name": "ፓስታ በስጎ", "price": 130, "category": "Fasting"},
    {"id": 2, "name": "ፓስታ በአትክልት", "price": 130, "category": "Fasting"},
    {"id": 3, "name": "ፍርፍር በቀይ", "price": 130, "category": "Fasting"},
    {"id": 4, "name": "ፍርፍር በአልጫ", "price": 130, "category": "Fasting"},
    {"id": 5, "name": "ሃፍ ሃፍ", "price": 160, "category": "Fasting"},
    {"id": 6, "name": "ሙሉ አገልግል", "price": 260, "category": "Fasting"},
    {"id": 7, "name": "ግማሽ አገልገል", "price": 180, "category": "Fasting"},
    {"id": 8, "name": "በየዓይነት", "price": 160, "category": "Fasting"},
    {"id": 9, "name": "ሙሉ ኮርኒስ", "price": 255, "category": "Fasting"},
    {"id": 10, "name": "ቴስቲ ወጥ", "price": 140, "category": "Fasting"},
    {"id": 11, "name": "ቴስቲ ለብለብ", "price": 150, "category": "Fasting"},
    {"id": 12, "name": "ቴስቲ ጥብስ", "price": 160, "category": "Fasting"},
    {"id": 13, "name": "ድብልቅ", "price": 240, "category": "Fasting"},
    {"id": 14, "name": "ሁሉ አይቅርብኝ", "price": 260, "category": "Fasting"},
    {"id": 15, "name": "ስጋ ፍርፍር", "price": 210, "category": "Non Fasting"},
    {"id": 16, "name": "ጥብስ ፍርፍር", "price": 260, "category": "Non Fasting"},
    {"id": 17, "name": "እንቁላል ፍርፍር", "price": 180, "category": "Non Fasting"},
    {"id": 18, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
    {"id": 19, "name": "ግማሽ ኮርኒስ (የፍስክ)", "price": 210, "category": "Non Fasting"},
    {"id": 20, "name": "ፓስታ በስጋ", "price": 210, "category": "Non Fasting"},
    {"id": 21, "name": "ስፔሻል ፍርፍር (የፍስክ)", "price": 310, "category": "Non Fasting"},
    {"id": 22, "name": "ሙሉ አገልግል (የፍስክ)", "price": 410, "category": "Non Fasting"},
    {"id": 23, "name": "ግማሽ አገልግል (የፍስክ)", "price": 260, "category": "Non Fasting"},
    {"id": 24, "name": "አይብ", "price": 260, "category": "Non Fasting"},
    {"id": 25, "name": "ስፔሻል ኮርኒስ (የጾም)", "price": 310, "category": "Specials"},
    {"id": 26, "name": "ጤና ገበታ (የጾም)", "price": 360, "category": "Specials"},
    {"id": 27, "name": "ጤና ስፔሻል ኮርኒስ (የፍስክ)", "price": 430, "category": "Specials"}
  ]
},
        {
        "telegram_id":825505972,
            "name": "Test Vendor",
            "menu": [
             {"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
                {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
                {"id": 78, "name": "Energy Drink", "price": 70, "category": "Drinks"},
                {"id": 79, "name": "Coca SD", "price": 75, "category": "Drinks"},
                {"id": 80, "name": "Sprite SD", "price": 75, "category": "Drinks"},
                {"id": 81, "name": "Fanta p.apple SD", "price": 75, "category": "Drinks"},
                {"id": 1, "name": "ሙሉ ኮርኒስ", "price": 210, "category": "Fasting"},
                {"id": 2, "name": "ሃፍ ሃፍ", "price": 160, "category": "Fasting"},
                {"id": 4, "name": "ሙሉ አገልግል", "price": 260, "category": "Fasting"},
                {"id": 5, "name": "ግማሽ አገልገል", "price": 160, "category": "Fasting"},
                {"id": 6, "name": "በየዓይነት", "price": 160, "category": "Fasting"},
                # {"id": 7, "name": "ተጋቢኖ", "price": 150, "category": "Fasting"},
                # {"id": 8, "name": "ፓስታ በስጎ", "price": 120, "category": "Fasting"},
                # {"id": 9, "name": "ፓስታ በአትክልት", "price": 120, "category": "Fasting"},
                # {"id": 10, "name": "ፓስታ በቴስቲ", "price": 120, "category": "Fasting"},
                # {"id": 11, "name": "ፍርፍር በቀይ", "price": 120, "category": "Fasting"},
                # {"id": 12, "name": "ፍርፍር በአልጫ", "price": 120, "category": "Fasting"},
                # {"id": 13, "name": "ስፔሻል ሽሮ", "price": 120, "category": "Fasting"},
                {"id": 14, "name": "ቴስቲ ወጥ", "price": 140, "category": "Fasting"},
                {"id": 15, "name": "ቴስቲ ለብለብ", "price": 150, "category": "Fasting"},
                {"id": 16, "name": "ቴስቲ ጥብስ", "price": 160, "category": "Fasting"},
                # {"id": 17, "name": "ቴስቲ ምንቸት", "price": 130, "category": "Fasting"},
                {"id": 18, "name": "ድብልቅ", "price": 240, "category": "Fasting"},
                {"id": 19, "name": "ፋሚሊ ኮምቦ", "price": 310, "category": "Fasting"},
                # {"id": 20, "name": "ጥብስ", "price": 250, "category": "Non Fasting"},
                # {"id": 21, "name": "ምንቸት", "price": 250, "category": "Non Fasting"},
                {"id": 22, "name": "ስጋ ፍርፍር", "price": 210, "category": "Non Fasting"},
                {"id": 23, "name": "ጥብስ ፍርፍር", "price": 230, "category": "Non Fasting"},
                {"id": 24, "name": "እንቁላል ፍርፍር", "price": 160, "category": "Non Fasting"},
                {"id": 25, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
                {"id": 27, "name": "ግማሽ ኮርኒስ", "price": 210, "category": "Non Fasting"},
                # {"id": 28, "name": "ፓስታ በእንቁላል", "price": 150, "category": "Non Fasting"},
                {"id": 29, "name": "ፓስታ በስጋ", "price": 190, "category": "Non Fasting"},
                {"id": 31, "name": "ስፔሻል ፍርፍር", "price": 260, "category": "Non Fasting"},
                # {"id": 32, "name": "እንቁላል በስጋ", "price": 200, "category": "Non Fasting"},
                {"id": 33, "name": "ሙሉ አገልግል (የፍስክ)", "price": 410, "category": "Non Fasting"},
                {"id": 34, "name": "ግማሽ አገልግል (የፍስክ)", "price": 260, "category": "Non Fasting"},
                # {"id": 35, "name": "ምስር በስጋ", "price": 180, "category": "Non Fasting"},
                # {"id": 36, "name": "ምስር በእንቁላል", "price": 150, "category": "Non Fasting"},
                {"id": 37, "name": "አይብ", "price": 210, "category": "Non Fasting"},
                {"id": 30, "name": "ስፔሻል ኮርኒስ (የጾም)", "price": 310, "category": "Specials"},
                {"id": 3, "name": "ጤና ገበታ (የጾም)", "price": 360, "category": "Specials"},
                {"id": 26, "name": "ጤና ስፔሻል ኮርኒስ (የፍስክ)", "price": 410, "category": "Specials"},
            ],
        },
        
        
       {
  "telegram_id": 577180165,
  "name": "ክርስቲና ምግብ ቤት",
  "menu": [
    {"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
    {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
    {"id": 78, "name": "Energy Drink", "price": 70, "category": "Drinks"},
    {"id": 79, "name": "Coca SD", "price": 75, "category": "Drinks"},
    {"id": 80, "name": "Sprite SD", "price": 75, "category": "Drinks"},
    {"id": 81, "name": "Fanta p.apple SD", "price": 75, "category": "Drinks"},
    {"id": 1, "name": "ፓስታ በስጎ", "price": 130, "category": "Fasting"},
    {"id": 2, "name": "ፓስታ በአትክልት", "price": 130, "category": "Fasting"},
    {"id": 3, "name": "ፓስታ በቴስቲ", "price": 130, "category": "Fasting"},
    {"id": 4, "name": "ፍርፍር በቀይ", "price": 130, "category": "Fasting"},
    {"id": 5, "name": "ፍርፍር በአልጫ", "price": 130, "category": "Fasting"},
    {"id": 6, "name": "ሃፍ ሃፍ", "price": 160, "category": "Fasting"},
    {"id": 7, "name": "ሙሉ አገልግል (የጾም)", "price": 290, "category": "Fasting"},
    {"id": 8, "name": "እርጥብ", "price": 130, "category": "Fasting"},
    {"id": 9, "name": "በየዓይነት", "price": 140, "category": "Fasting"},
    {"id": 10, "name": "ሙሉ ኮርኒስ (የጾም)", "price": 260, "category": "Fasting"},
    {"id": 11, "name": "ቲማቲም ለብለብ", "price": 130, "category": "Fasting"},
    {"id": 12, "name": "ተጋቢኖ", "price": 150, "category": "Fasting"},
    {"id": 13, "name": "ቴስቲ ወጥ", "price": 130, "category": "Fasting"},
    {"id": 14, "name": "ቴስቲ ለብለብ", "price": 150, "category": "Fasting"},
    {"id": 15, "name": "ቴስቲ ጥብስ", "price": 140, "category": "Fasting"},
    {"id": 16, "name": "ቤቴን አስረሳሽኝ (የጾም)", "price": 180, "category": "Fasting"},
    {"id": 17, "name": "ሁሉ አይቅርብኝ (የጾም)", "price": 290, "category": "Fasting"},
    {"id": 18, "name": "ስጋ ፍርፍር", "price": 190, "category": "Non Fasting"},
    {"id": 19, "name": "ጥብስ ፍርፍር", "price": 260, "category": "Non Fasting"},
    {"id": 20, "name": "እንቁላል ፍርፍር", "price": 160, "category": "Non Fasting"},
    {"id": 21, "name": "እንቁላል በስጋ", "price": 210, "category": "Non Fasting"},
    {"id": 22, "name": "ሙሉ ኮርኒስ (የፍስክ)", "price": 360, "category": "Non Fasting"},
    {"id": 23, "name": "ፓስታ በስጋ", "price": 190, "category": "Non Fasting"},
    {"id": 24, "name": "ስፔሻል ፍርፍር (የፍስክ)", "price": 310, "category": "Non Fasting"},
    {"id": 25, "name": "ሙሉ አገልግል (የፍስክ)", "price": 410, "category": "Non Fasting"},
    {"id": 26, "name": "ፍርፍር በቅቤ", "price": 160, "category": "Non Fasting"},
    {"id": 27, "name": "አይብ", "price": 210, "category": "Non Fasting"},
    {"id": 28, "name": "ስፔሻል ኮርኒስ (የጾም)", "price": 310, "category": "Specials"},
    {"id": 29, "name": "ምንትዋብ", "price": 210, "category": "Specials"},
    {"id": 30, "name": "ይምጡ በዝና", "price": 260, "category": "Specials"}
  ]
},

        {
        "telegram_id": 5771801656,
            "name": "መቅዲ ምግብ ቤት",
            "menu": [
                 {"id": 77, "name": "ዳቦ", "price": 15, "category": "Extras"},
                {"id": 78, "name": "ግማሽ ሊትር ውሃ", "price": 20, "category": "Extras"},
                {"id": 79, "name": "1 ሊትር ውሃ", "price": 30, "category": "Extras"},
                {"id": 1, "name": "ፓስታ በስጎ", "price": 130, "category": "Fasting"},
                {"id": 2, "name": "ፓስታ በአትክልት", "price": 130, "category": "Fasting"},
                {"id": 3, "name": "ፓስታ በቴስቲ", "price": 130, "category": "Fasting"},
                {"id": 4, "name": "ፍርፍር በቀይ", "price": 130, "category": "Fasting"},
                {"id": 5, "name": "ፍርፍር በአልጫ", "price": 130, "category": "Fasting"},
                {"id": 6, "name": "ሃፍ ሃፍ", "price": 150, "category": "Fasting"},
                {"id": 7, "name": "ሙሉ አገልግል (የጾም)", "price": 300, "category": "Fasting"},
                {"id": 8, "name": "ሹፌር ማኛ", "price": 300, "category": "Fasting"},
                {"id": 9, "name": "በየዓይነት", "price": 300, "category": "Fasting"},
                {"id": 10, "name": "ሙሉ ኮርኒስ (የጾም)", "price": 250, "category": "Fasting"},
                {"id": 11, "name": "ቲማቲም ለብለብ", "price": 130, "category": "Fasting"},
                {"id": 12, "name": "ተጋቢኖ", "price": 150, "category": "Fasting"},
                {"id": 13, "name": "ቴስቲ ወጥ", "price": 130, "category": "Fasting"},
                {"id": 14, "name": "ቴስቲ ጥብስ", "price": 130, "category": "Fasting"},
                {"id": 15, "name": "ግማሽ ኮርኒስ (የጾም)", "price": 150, "category": "Fasting"},
                # {"id": 17, "name": "ቴስቲ ምንቸት", "price": 130, "category": "Fasting"},
                {"id": 16, "name": "ቤቴን አስረሳሽኝ (የጾም)", "price": 200, "category": "Fasting"},
                {"id": 17, "name": "ሁሉ አይቅርብኝ (የጾም)", "price": 350, "category": "Fasting"},
                # {"id": 20, "name": "ጥብስ", "price": 250, "category": "Non Fasting"},
                # {"id": 21, "name": "ምንቸት", "price": 250, "category": "Non Fasting"},
                {"id": 18, "name": "ስጋ ፍርፍር", "price": 200, "category": "Non Fasting"},
                {"id": 19, "name": "ስፔሻል ፍርፍር", "price": 250, "category": "Non Fasting"},
                {"id": 20, "name": "እንቁላል ፍርፍር", "price": 150, "category": "Non Fasting"},
                {"id": 21, "name": "እንቁላል በስጋ", "price": 200, "category": "Non Fasting"},
                {"id": 22, "name": "ሙሉ ኮርኒስ (የፍስክ)", "price": 350, "category": "Non Fasting"},
                {"id": 23, "name": "ግማሽ ኮርኒስ(የፍስክ)", "price": 220, "category": "Non Fasting"},
                {"id": 24, "name": "ፓስታ በእንቁላል", "price": 150, "category": "Non Fasting"},
                {"id": 25, "name": "ፓስታ በስጋ", "price": 200, "category": "Non Fasting"},
                # {"id": 32, "name": "እንቁላል በስጋ", "price": 200, "category": "Non Fasting"},
                {"id": 26, "name": "ሙሉ አገልግል (የፍስክ)", "price": 400, "category": "Non Fasting"},
                {"id": 27, "name": "ፍርፍር በቅቤ", "price": 150, "category": "Non Fasting"},
                # {"id": 35, "name": "ምስር በስጋ", "price": 180, "category": "Non Fasting"},
                # {"id": 36, "name": "ምስር በእንቁላል", "price": 150, "category": "Non Fasting"},
                {"id": 28, "name": "አይብ", "price": 200, "category": "Non Fasting"},
                {"id": 29, "name": "ቦዘና ሽሮ", "price": 200, "category": "Specials"},
                {"id": 30, "name": "ምንትዋብ", "price": 200, "category": "Specials"},
                {"id": 31, "name": "ይምጡ በዝና", "price": 350, "category": "Specials"},
            ],
        },
        
    ]
    
    

    # async with db._open_connection() as conn:
    # # Delete all existing vendors
    #     # await conn.execute("TRUNCATE TABLE vendors RESTART IDENTITY CASCADE")

    #     # Now insert fresh seed data
    #     for v in vendors:
    #         menu_json = json.dumps(v["menu"])
    #         await conn.execute(
    #             """
    #             INSERT INTO vendors (telegram_id, name, menu_json)
    #             VALUES ($1, $2, $3)
    #             """,
    #             v["telegram_id"], v["name"], menu_json
    #         )
    # print("✅ Vendors table truncated and seeded successfully")
    
    
    async with db._open_connection() as conn:
        for v in vendors:
            menu_json = json.dumps(v["menu"])
            await conn.execute(
                """
                INSERT INTO vendors (telegram_id, name, menu_json)
                VALUES ($1, $2, $3)
                ON CONFLICT (telegram_id)
                DO UPDATE SET
                    name = EXCLUDED.name,
                    menu_json = EXCLUDED.menu_json,
                    updated_at = CURRENT_TIMESTAMP
                """,
                v["telegram_id"], v["name"], menu_json
            )
    print("✅ Vendors table upserted successfully")




async def seed_delivery_guys(db: Database) -> None:
    """
    Reset and seed the delivery_guys table with sample entries.
    """

    # Each tuple: (user_id, telegram_id, name, campus, active,
    #              total_deliveries, accepted_requests, total_requests,
    #              coins, xp, level)
    delivery_guys_data: List[Tuple] = [
    (1001, settings.DG_IDS["Dagmawi"], "Dagmawi", "6kilo", True, 12, 14, 16, 10, 260, 3),
    (1002, settings.DG_IDS["Muktar"], "Muktar", "5kilo", True, 8, 10, 12, 80, 150, 2),
    (1003, settings.DG_IDS["Yonatan"], "Yonatan", "6kilo", True, 5, 6, 8, 50, 100, 1),
    # Random demo entries
    (random.randint(100000000, 999999999), random.randint(100000000, 999999999),
     "Selam", "6kilo", False, 15, 18, 20, 150, 300, 4),
    (random.randint(100000000, 999999999), random.randint(100000000, 999999999),
     "Kebede", "6kilo", False, 20, 22, 24, 200, 400, 5),
]



    async with db._open_connection() as conn:
        # Clear table first
        await conn.execute("TRUNCATE delivery_guys RESTART IDENTITY CASCADE")

        insert_sql = """
            INSERT INTO delivery_guys 
            (user_id, telegram_id, name, campus, active,
             total_deliveries, accepted_requests, total_requests,
             coins, xp, level)
            VALUES ($1::BIGINT, $2::BIGINT, $3, $4, $5,
                    $6, $7, $8, $9, $10, $11)
        """

        for row in delivery_guys_data:
            await conn.execute(insert_sql, *row)

    print("✅ delivery_guys table reset and seeded with entries")
    
async def seed_specific_dg(db: Database) -> None:
    """
    Inspect and then delete a specific delivery guy and their user record by telegram_id.
    """
    telegram_id = 1701238322  # the DG you want to remove

    async with db._open_connection() as conn:
        # Look up the user record first
        user_row = await conn.fetchrow(
            "SELECT id, telegram_id, role, first_name, phone, campus FROM users WHERE telegram_id = $1",
            telegram_id
        )
        if user_row:
            print("🔎 Found user before delete:", dict(user_row))
        else:
            print("⚠️ No user found with telegram_id", telegram_id)

        # Look up the delivery_guy record too
        dg_row = await conn.fetchrow(
            "SELECT id, user_id, telegram_id, name, campus, phone, active FROM delivery_guys WHERE telegram_id = $1",
            telegram_id
        )
        if dg_row:
            print("🔎 Found delivery_guy before delete:", dict(dg_row))
        else:
            print("⚠️ No delivery_guy found with telegram_id", telegram_id)

        # await conn.execute("DELETE FROM delivery_guys WHERE telegram_id = $1", telegram_id)
        # await conn.execute("DELETE FROM users WHERE telegram_id = $1", telegram_id)

    print("✅ Deleted specific delivery guy with telegram_id", telegram_id)



async def generate_delivery_guy_row(db: Database, telegram_id: int) -> int:
    """
    Ensure a delivery_guy row exists for the given telegram_id.
    If missing, create it from the corresponding users record.
    Returns the delivery_guy.id.
    """
    async with db._open_connection() as conn:
        # 1. Fetch the user record
        user_row = await conn.fetchrow(
            "SELECT id, first_name, campus, phone FROM users WHERE telegram_id = $1",
            telegram_id
        )
        if not user_row:
            raise ValueError(f"No user found with telegram_id={telegram_id}")

        user_id = user_row["id"]
        name = user_row["first_name"]
        campus = user_row["campus"]
        phone = user_row["phone"]

        # 2. Try to insert delivery_guy row (or update if exists)
        dg_id = await conn.fetchval(
            """
            INSERT INTO delivery_guys (user_id, telegram_id, name, campus, phone, active, total_deliveries)
            VALUES ($1, $2, $3, $4, $5, TRUE, 0)
            ON CONFLICT (user_id) DO UPDATE
            SET telegram_id = EXCLUDED.telegram_id,
                name = EXCLUDED.name,
                campus = EXCLUDED.campus,
                phone = EXCLUDED.phone,
                active = TRUE
            RETURNING id
            """,
            user_id, telegram_id, name, campus, phone
        )

    print(f"✅ Delivery guy row ensured for telegram_id {telegram_id}, id={dg_id}")
    return int(dg_id)




async def debug_list_delivery_guys(db):
    rows = await db._pool.fetch("SELECT * FROM delivery_guys ORDER BY id")
    for r in rows:
        print(dict(r))




async def update_menu_item_price(
    db: Database,
    vendor_telegram_id: int,
    item_id: int,
    new_price: int,
) -> None:
    async with db._open_connection() as conn:
        await conn.execute(
            """
            UPDATE vendors
            SET menu_json = (
                SELECT jsonb_agg(
                    CASE
                        WHEN (item->>'id')::int = $2
                        THEN jsonb_set(item, '{price}', to_jsonb($3::int), false)
                        ELSE item
                    END
                )
                FROM jsonb_array_elements(menu_json::jsonb) AS item
            )
            WHERE telegram_id = $1
            """,
            vendor_telegram_id,
            item_id,
            new_price,
        )

    print("✅ Menu item price updated successfully")


import json
from typing import Dict, Any

async def replace_menu_item(
    db: Database,
    vendor_telegram_id: int,
    item: Dict[str, Any],
) -> None:
    """
    Remove any existing menu item with the same id and insert the new one.
    Ensures menu_json stays a valid JSON array of objects.
    """
    async with db._open_connection() as conn:
        await conn.execute(
            """
            UPDATE vendors
            SET menu_json = (
                SELECT jsonb_agg(elem)
                FROM (
                    -- keep all items except the one with matching id
                    SELECT elem
                    FROM jsonb_array_elements(menu_json::jsonb) AS elem
                    WHERE (elem->>'id')::int <> $2

                    UNION ALL
                    -- add the new corrected item
                    SELECT to_jsonb($3::jsonb)
                ) AS elem
            )
            WHERE telegram_id = $1
            """,
            vendor_telegram_id,
            item["id"],              # the id to replace
            json.dumps(item),        # the new item as JSON
        )
    print(f"✅ Replaced menu item {item['id']} for vendor {vendor_telegram_id}")
